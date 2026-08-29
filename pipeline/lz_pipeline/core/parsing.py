"""Workbook -> spec dict (sentinel-driven sheet/table parser)."""

import json
from pathlib import Path
from openpyxl import load_workbook
from .helpers import _truthy, _coerce


def parse_workbook(path: Path) -> dict:
    """Return {sheet_name: {table_name: parsed_table}}.

    parsed_table shape depends on kind, inferred from headers we encounter:
      - scalar       -> dict {field_name: value_or_None}
      - list-single  -> list[str]
      - object-table -> list[dict]; rows without Enabled=True are filtered out
                        if Enabled column is present.
    """
    from lz_spec.schema import INFO_SHEETS
    wb = load_workbook(path, data_only=True)
    out = {}
    for ws in wb.worksheets:
        if ws.title in INFO_SHEETS:
            continue  # informational
        out[ws.title] = _parse_sheet(ws)
    return out


def _parse_sheet(ws) -> dict:
    rows = list(ws.iter_rows(values_only=True))
    tables = {}
    i = 0
    while i < len(rows):
        row = rows[i]
        first = row[0]
        if isinstance(first, str) and first.startswith("### "):
            name = first[4:].strip()
            # Skip the optional description row
            j = i + 1
            # Skip empty / single-cell description rows until headers
            while j < len(rows) and _is_blank_or_text_only(rows[j]):
                j += 1
            if j >= len(rows):
                break
            headers = [str(c).strip() if c is not None else "" for c in rows[j]]
            data_start = j + 1
            # Determine table kind from headers
            if headers[:6] == ["Field", "Type", "Default", "Sample", "Description", "Value"]:
                # scalar
                data_end, parsed = _parse_scalar(rows, data_start)
                tables[name] = parsed
                i = data_end
            elif headers[0] == "Value" and (len(headers) == 1 or headers[1] in ("Sample", "")):
                # list-single
                data_end, parsed = _parse_list_single(rows, data_start)
                tables[name] = parsed
                i = data_end
            else:
                # object-table; possibly skip a "(type)" annotation row
                if data_start < len(rows) and _looks_like_type_row(rows[data_start]):
                    data_start += 1
                data_end, parsed = _parse_object_table(rows, data_start, headers)
                tables[name] = parsed
                i = data_end
        else:
            i += 1
    return tables


def _is_blank_or_text_only(row) -> bool:
    nonblank = [c for c in row if c is not None and str(c).strip() != ""]
    if not nonblank:
        return True
    # description row: only first cell has content
    return len(nonblank) == 1


def _looks_like_type_row(row) -> bool:
    cells = [c for c in row if c is not None]
    if not cells:
        return False
    return all(isinstance(c, str) and c.startswith("(") and c.endswith(")") for c in cells)


def _parse_scalar(rows, start):
    result = {}
    i = start
    while i < len(rows):
        row = rows[i]
        if row[0] is None or (isinstance(row[0], str) and row[0].startswith("### ")):
            break
        field = row[0]
        typ = row[1]
        value_raw = row[5] if len(row) > 5 else None
        if isinstance(field, str) and field.strip():
            if value_raw is None or (isinstance(value_raw, str) and value_raw.strip() == ""):
                result[field.strip()] = None  # signal "use module default"
            else:
                result[field.strip()] = _coerce(value_raw, typ)
        i += 1
    return i, result


def _parse_list_single(rows, start):
    result = []
    i = start
    blanks = 0
    while i < len(rows):
        row = rows[i]
        if row[0] is not None and isinstance(row[0], str) and row[0].startswith("### "):
            break
        v = row[0]
        if v is None or (isinstance(v, str) and v.strip() == ""):
            blanks += 1
            if blanks >= 3:
                break
        else:
            blanks = 0
            result.append(str(v).strip())
        i += 1
    return i, result


def _parse_object_table(rows, start, headers):
    result = []
    i = start
    blanks = 0
    while i < len(rows):
        row = rows[i]
        if row and row[0] is not None and isinstance(row[0], str) and row[0].startswith("### "):
            break
        # Check if row is entirely blank
        nonblank = any(c is not None and str(c).strip() != "" for c in row)
        if not nonblank:
            blanks += 1
            if blanks >= 3:
                break
            i += 1
            continue
        blanks = 0
        d = {}
        for col_idx, h in enumerate(headers):
            if not h:
                continue
            v = row[col_idx] if col_idx < len(row) else None
            if v is None or (isinstance(v, str) and v.strip() == ""):
                d[h] = None
            else:
                d[h] = v
        # Filter on Enabled if column present
        if "Enabled" in d:
            if not _truthy(d["Enabled"]):
                i += 1
                continue
        result.append(d)
        i += 1
    return i, result
