"""Generate the deployed-resource checklist workbook from pulled states.

Customer-agnostic: the provider-alias -> account mapping is parsed from each
env's provider files (generated + static), not hardcoded. Envs without a
state file (or with an empty state) land on the Pending sheet.

Usage:
    py tools/gen_checklist.py --envs-dir ..\\huawei-lz\\envs-acme \
        --states-dir <dir with state-<env>.json> --out checklist.xlsx \
        [--title "Example Landing Zone Resource Checklist"] [--master-label "Master (management account)"]

Refresh states first:  cd envs\\<env> && terraform state pull > state-<env>.json
"""

import argparse
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from envtree import (BOX, HDR_FILL, HDR_FONT, WRAP,
                     env_dirs, state, alias_accounts, provider_alias_of)

# Generic, platform-level descriptions (no customer specifics).
DESC = {
    "huaweicloud_organizations_organization": "The organization itself; the root under which all member accounts and OUs live.",
    "huaweicloud_organizations_organizational_unit": "Organizational units grouping member accounts for policy attachment.",
    "huaweicloud_organizations_account": "Member accounts created in the organization, each with its cross-account access agency.",
    "huaweicloud_organizations_trusted_service": "Org-integrated services (CTS, Config, LTS, RAM and so on) enabled organization-wide.",
    "huaweicloud_organizations_delegated_administrator": "Hands a service's org-wide administration to a member account.",
    "huaweicloud_organizations_policy": "Service control policy documents (identity guardrails, tag enforcement).",
    "huaweicloud_organizations_policy_attach": "Attaches the SCP documents to the organization root or OUs, making them effective.",
    "huaweicloud_identitycenter_instance": "IAM Identity Center instance providing workforce SSO.",
    "huaweicloud_identitycenter_registered_region": "Registers the home region with Identity Center before the service starts.",
    "huaweicloud_identitycenter_permission_set": "Permission sets that users assume in member accounts.",
    "huaweicloud_identitycenter_system_policy_attachment": "Binds Huawei system policies to the permission sets.",
    "huaweicloud_identitycenter_mfa_management_setting": "Identity Center MFA behaviour setting.",
    "huaweicloud_identitycenter_password_policy": "Password policy for Identity Center local users.",
    "huaweicloud_identity_password_policy": "Per-account IAM password policy baseline (length, complexity, expiry).",
    "huaweicloud_identity_login_policy": "Per-account IAM login protection baseline (lockout, session limits).",
    "huaweicloud_identity_protection_policy": "Per-account operation-protection baseline for sensitive console actions.",
    "huaweicloud_identity_trust_agency": "Service agencies allowing platform services to act within the account.",
    "huaweicloud_enterprise_project": "Cost-center enterprise projects used to group and bill resources.",
    "huaweicloud_enterprise_project_authority": "Enables the enterprise project capability in an account.",
    "huaweicloud_tms_tags": "Predefined tag dictionary fanned out to every account.",
    "huaweicloud_vpc": "Hub and spoke VPCs.",
    "huaweicloud_vpc_subnet": "Subnets within the hub and spoke VPCs.",
    "huaweicloud_vpc_route": "VPC route-table entries sending traffic to the Enterprise Router or NAT.",
    "huaweicloud_vpc_flow_log": "Per-VPC flow logs recording traffic into dedicated LTS groups.",
    "huaweicloud_networking_secgroup": "Baseline security groups for hub and spoke workloads.",
    "huaweicloud_networking_secgroup_rule": "The rules inside the baseline security groups.",
    "huaweicloud_er_instance": "The hub Enterprise Router - the routing core connecting all VPCs and the VPN.",
    "huaweicloud_er_route_table": "ER route tables implementing the inspection topology.",
    "huaweicloud_er_vpc_attachment": "Connects each hub/spoke VPC to the Enterprise Router.",
    "huaweicloud_er_association": "Associates each attachment to its route table.",
    "huaweicloud_er_propagation": "Propagates each attachment's routes into its route table.",
    "huaweicloud_er_static_route": "Steering routes forcing traffic through the firewall and NAT.",
    "huaweicloud_er_flow_log": "Flow log on the Enterprise Router for routing-level visibility.",
    "huaweicloud_cfw_firewall": "The Cloud Firewall instance inspecting east-west and border traffic.",
    "huaweicloud_cfw_lts_log": "Streams firewall traffic/access/attack logs into LTS.",
    "huaweicloud_cfw_address_group": "Named IP sets used by firewall rules.",
    "huaweicloud_cfw_address_group_member": "The individual CIDR entries inside the address groups.",
    "huaweicloud_cfw_domain_name_group": "Named domain sets for egress whitelisting.",
    "huaweicloud_cfw_service_group": "Named port sets used by firewall rules.",
    "huaweicloud_cfw_service_group_member": "The protocol/port entries inside the service groups.",
    "huaweicloud_cfw_acl_rule": "Firewall rules, including the bottom-pinned catch-all denies.",
    "huaweicloud_natv3_gateway": "Public NAT gateways for centralized internet egress.",
    "huaweicloud_vpn_gateway": "Site-to-cloud VPN gateway.",
    "huaweicloud_vpn_customer_gateway": "Far-end gateway instances registered as VPN peers.",
    "huaweicloud_vpn_connection": "IPsec connections between the VPN gateway and customer gateways.",
    "huaweicloud_dns_endpoint": "Hybrid DNS resolver endpoints (inbound/outbound).",
    "huaweicloud_dns_resolver_rule": "Forwarding rules sending corporate domains to on-premises DNS.",
    "huaweicloud_dns_resolver_rule_associate": "Attaches each forwarding rule to its VPCs.",
    "huaweicloud_dns_resolver_access_log": "DNS query logging into LTS.",
    "huaweicloud_dns_zone": "Hosted DNS zones (public or private).",
    "huaweicloud_dns_recordset": "Record sets within the hosted zones.",
    "huaweicloud_cts_tracker": "Audit trackers: the org-wide tracker plus minimal per-account trackers.",
    "huaweicloud_obs_bucket": "OBS buckets (state, audit trail, log archive, recorder delivery).",
    "huaweicloud_obs_bucket_bpa": "Blocks all public access on the buckets.",
    "huaweicloud_obs_bucket_policy": "Bucket policies (TLS-only, cross-account write).",
    "huaweicloud_kms_key": "KMS keys encrypting the audit and archive buckets.",
    "huaweicloud_lts_group": "LTS log groups.",
    "huaweicloud_lts_stream": "The log streams inside those groups.",
    "huaweicloud_lts_transfer": "Scheduled transfers moving logs into the archive bucket.",
    "huaweicloud_lts_log_converge_switch": "Enables the log-admin account to receive logs from other accounts.",
    "huaweicloud_lts_log_converge": "Per-account log convergence mappings feeding central aggregation.",
    "huaweicloud_rms_resource_recorder": "Config resource recorders capturing resource inventory and changes.",
    "huaweicloud_rms_resource_aggregator": "Organization-wide Config aggregator.",
    "huaweicloud_rms_organizational_assignment_package": "Org-wide conformance packages evaluating compliance continuously.",
    "huaweicloud_smn_topic": "SMN notification topics for alarms and operational alerts.",
    "huaweicloud_smn_subscription": "Subscriptions on those topics.",
    "huaweicloud_ces_one_click_alarm": "Cloud Eye one-click alarm bundles per account.",
    "huaweicloud_ram_organization": "Enables organization-level resource sharing (RAM).",
    "huaweicloud_ram_resource_share": "Shares the hub Enterprise Router with the workload accounts.",
    "huaweicloud_secmaster_workspace": "SecMaster security-operations workspace.",
    "huaweicloud_waf_dedicated_instance": "Dedicated WAF instance.",
    "huaweicloud_waf_policy": "WAF protection policy.",
    "huaweicloud_waf_dedicated_domain": "Domains protected by the dedicated WAF.",
    "huaweicloud_antiddos_basic": "Anti-DDoS traffic-cleaning thresholds on public EIPs.",
    "time_sleep": "Terraform-internal wait timers for propagation. No cloud resource.",
}

ENV_FILL = PatternFill("solid", fgColor="DDEBF7")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--envs-dir", required=True)
    ap.add_argument("--states-dir", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--title", default="Landing Zone - deployed resource checklist")
    ap.add_argument("--master-label", default="Master (management account)")
    args = ap.parse_args()

    envs_dir = Path(args.envs_dir)
    states_dir = Path(args.states_dir)

    rows = defaultdict(int)   # (env, account, type) -> count
    pending = []
    envs = [d.name for d in env_dirs(envs_dir)]
    for env in envs:
        st = state(states_dir, env)
        aliases = alias_accounts(envs_dir, env)
        managed = [r for r in st.get("resources", []) if r.get("mode") == "managed" and r.get("instances")]
        if not managed:
            reason = "no pulled state" if not st else "state is empty (never applied)"
            pending.append((env, reason))
            continue
        for r in managed:
            alias = provider_alias_of(r.get("provider", ""))
            acct = aliases.get(alias, args.master_label if not alias else alias)
            rows[(env, acct, r["type"])] += len(r.get("instances", []))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Deployed Resources"
    ws["A1"] = args.title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = ("Source: terraform state of every environment. One row per resource type per "
                "account per environment. Count = deployed instances.")
    ws["A2"].font = Font(size=9, italic=True, color="595959")

    r = 4
    for c, v in enumerate(["Environment", "Account", "Resource", "Count", "What it does"], 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font, cell.fill, cell.border = HDR_FONT, HDR_FILL, BOX
    ws.freeze_panes = "A5"
    r += 1

    for env in envs:
        env_rows = sorted((k, v) for k, v in rows.items() if k[0] == env)
        if not env_rows:
            continue
        first = True
        for (e, acct, typ), n in sorted(env_rows,
                                        key=lambda kv: (kv[0][1] != args.master_label, kv[0][1], kv[0][2])):
            vals = [env if first else "", acct, typ.replace("huaweicloud_", ""), n, DESC.get(typ, "")]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border, cell.alignment = BOX, WRAP
                if first:
                    cell.fill = ENV_FILL
            first = False
            r += 1
    for i, w in enumerate([16, 28, 40, 7, 78], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    p = wb.create_sheet("Pending (not yet applied)")
    p["A1"] = "Environments without deployed state"
    p["A1"].font = Font(bold=True, size=12)
    for c, v in enumerate(["Environment", "Status"], 1):
        cell = p.cell(row=3, column=c, value=v)
        cell.font, cell.fill, cell.border = HDR_FONT, HDR_FILL, BOX
    for i, (env, reason) in enumerate(pending, 4):
        for c, v in enumerate([env, reason], 1):
            cell = p.cell(row=i, column=c, value=v)
            cell.border, cell.alignment = BOX, WRAP
    for i, w in enumerate([20, 60], 1):
        p.column_dimensions[get_column_letter(i)].width = w

    wb.save(args.out)
    print(f"written: {args.out}")
    print(f"rows: {len(rows)} (env,account,type) | instances: {sum(rows.values())} | pending: {len(pending)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
