"""Dump a filled LZ Assessment Questionnaire to JSON for conversion.

Mechanical extraction only - no interpretation. The /questionnaire-to-spec
skill consumes this dump: prose answers are interpreted by the agent, appendix
rows are copied VERBATIM into the draft spec (never retyped).

Usage: py tools/dump_questionnaire.py <filled.xlsx> [-o out.json]
"""

import argparse
import json
import re
import sys
from pathlib import Path

SURVEY_SHEETS = ["Core Questions", "Deep-Dive Questions"]

# Deterministic pre-model redaction: a value a customer types after a secret
# designator ("the PSK is ...", "password: ...") never reaches the agent -
# it is replaced with a sentinel BEFORE the dump is written. Column-name
# matching covers appendix tables. The model cannot echo what it never sees;
# the eval secret canary covers the residual (secrets pasted with no
# designator, which no deterministic rule can recognize).
SECRET_SENTINEL = "[SECRET-REDACTED]"
_SECRET_VALUE = re.compile(
    r"(?i)\b(psk|pre-?shared\s+key|password|passphrase|secret(?:\s+key)?|"
    r"access\s+key|api\s+key|token|ak/sk)\b\s*(?:is|was|[:=])\s*"
    r"[\"']?([^\s\"',;]+)")
_SECRET_COLUMN = re.compile(r"(?i)\b(psk|password|passphrase|secret|token)\b")


def _redact_answer(text: str):
    """(redacted_text, n_redactions) - values after secret designators only."""
    out, n = [], 0
    last = 0
    for m in _SECRET_VALUE.finditer(text):
        out.append(text[last:m.start(2)])
        out.append(SECRET_SENTINEL)
        last = m.end(2)
        n += 1
    out.append(text[last:])
    return "".join(out), n


def dump(path: Path) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)

    wiring = {}
    meta = {}
    if "_wiring" in wb.sheetnames:
        for row in wb["_wiring"].iter_rows(min_row=2, values_only=True):
            ref, tier, cat, targets, default = (list(row) + [None] * 5)[:5]
            if ref == "_meta":
                for part in (targets or "").split(";"):
                    if "=" in part:
                        k, v = part.split("=", 1)
                        meta[k.strip()] = v.strip()
            elif ref:
                wiring[str(ref)] = {
                    "targets": [t.strip() for t in (targets or "").split(";") if t.strip()],
                    "default_if_silent": default or "",
                }

    answers = []
    for sheet in SURVEY_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        for row in wb[sheet].iter_rows(min_row=3):
            ref, question, guidance, response = [row[i].value if i < len(row) else None
                                                 for i in range(4)]
            ref = str(ref).strip() if ref else ""
            if not re.fullmatch(r"[CD]\d+", ref):   # category band / blank row
                continue
            w = wiring.get(ref, {})
            raw = str(response).strip() if response is not None else ""
            answer, n_secrets = _redact_answer(raw)
            entry = {
                "ref": ref,
                "question": (question or "").strip(),
                "answer": answer,
                "targets": w.get("targets", []),
                "default_if_silent": w.get("default_if_silent", ""),
            }
            if n_secrets:
                entry["secret_present"] = True
                entry["secret_note"] = (f"{n_secrets} secret value(s) redacted at "
                                        "intake; collect out-of-band into a secret "
                                        "store and reference it from the spec")
            answers.append(entry)

    appendices = {}
    for sheet in wb.sheetnames:
        if not sheet.startswith("Appendix"):
            continue
        ws = wb[sheet]
        headers = [c.value for c in ws[3] if c.value is not None]
        rows = []
        secret_cols = [h for h in headers if _SECRET_COLUMN.search(str(h))]
        for row in ws.iter_rows(min_row=4, values_only=True):
            vals = [("" if v is None else str(v).strip()) for v in row[:len(headers)]]
            if not any(vals):
                continue
            if vals[0].startswith("(example)"):
                continue
            r = dict(zip(headers, vals))
            for h in secret_cols:
                if r.get(h):
                    r[h] = SECRET_SENTINEL
            rows.append(r)
        ref = sheet.split(" ")[1]     # "Appendix A - Accounts" -> "A"
        w = wiring.get(ref, {})
        appendices[ref] = {"sheet": sheet, "targets": w.get("targets", []), "rows": rows}

    return {"source_file": path.name, "meta": meta,
            "answers": answers, "appendices": appendices}


def main(argv=None):
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="filled questionnaire")
    ap.add_argument("-o", "--out", help="output JSON path (default: stdout)")
    args = ap.parse_args(argv)

    data = dump(Path(args.xlsx))
    answered = sum(1 for a in data["answers"] if a["answer"])
    text = json.dumps(data, indent=2, ensure_ascii=False)
    if args.out:
        Path(args.out).write_text(text, encoding="utf-8")
        print(f"wrote {args.out}: {answered}/{len(data['answers'])} answered, "
              f"{sum(len(a['rows']) for a in data['appendices'].values())} appendix rows")
    else:
        print(text)
    return 0


if __name__ == "__main__":
    sys.exit(main())
