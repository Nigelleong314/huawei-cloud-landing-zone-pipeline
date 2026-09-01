"""Generate the customer-facing LZ Assessment Questionnaire FROM the schema.

Pre-engagement discovery instrument: open-ended questions a customer answers
alone, before any workshop. A filled copy is converted to a draft
lz.spec.<customer>.json by the /questionnaire-to-spec skill (via
tools/dump_questionnaire.py) — gaps become the decisions-needed list.

The question CATALOGUE below is the source of truth; the xlsx is a rendering.
Each question carries wiring: the "Sheet.Table[.field]" targets it informs.
A coverage check walks lz_spec/schema.py and fails the build if any
non-exempt table is unreachable from every question — so schema drift
surfaces here, not in a stale questionnaire.

Usage: py tools/gen_questionnaire.py [-o out.xlsx] [--check]
"""

import argparse
import datetime
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent.parent   # workspace
sys.path.insert(0, str(ROOT))
sys.path.insert(0, str(ROOT / "lz_spec"))

QUESTIONNAIRE_VERSION = "1.0"
DEFAULT_OUT = ROOT / "HuaweiCloud-LZ-Assessment-Questionnaire.xlsx"

# Tables no customer question should cover: engineer-derived plumbing, or
# reserved/not-implemented. Everything else in schema.SHEETS must be wired.
ENGINEER_ONLY = {
    "01_Foundation.EnabledPolicyTypes",   # fixed org policy-type enum
    "01_Foundation.TrustedServices",      # baseline org integrations
    "03_Identity.ServiceAgencies",        # service-to-service trust plumbing
    "05_Network.ERAvailabilityZones",     # derived from region
    "05_Network.HubERAttachments",        # derived from hub topology
    "05_Network.SpokeERAttachments",      # derived from spoke topology
    "05_Network.RAMSharePrincipals",      # derived from account list
    "06_Observability.LogConverge",       # schema says: leave empty, derived
    "11_SGACL.NetworkACLs",               # RESERVED (LZR-031)
    "11_SGACL.ACLRules",                  # RESERVED (LZR-031)
}

CATEGORIES = [
    "Organization & Accounts",
    "Identity & Access",
    "Network",
    "Security",
    "Compliance & Audit",
    "Operations & Monitoring",
    "Finance & Cost Management",
]

# (tier, category, question, guidance, wiring, default_if_silent)
#   tier: "core" | "deep"
#   wiring: list of "Sheet.Table" or "Sheet.Table.field" targets (may be
#           empty for pure-context questions)
Q = []


def _q(tier, cat, question, guidance, wiring, default="", example=""):
    assert cat in CATEGORIES, cat
    Q.append({"tier": tier, "category": cat, "question": question,
              "guidance": guidance, "wiring": wiring, "default": default,
              "example": example})


# ── Organization & Accounts ─────────────────────────────────────────────────
_q("core", "Organization & Accounts",
   "Describe your current cloud footprint: which providers you use (AWS, Azure, GCP, Alibaba, Huawei), roughly how many accounts / subscriptions / projects, and what each is for.",
   "Free text; attach an account inventory or org chart if you have one. This shapes the account and OU structure we propose.",
   ["01_Foundation.OrganizationalUnits"],
   example="AWS: ~25 accounts, one per team; Azure: 8 subscriptions (legacy ERP); no Huawei footprint yet.")
_q("core", "Organization & Accounts",
   "If you already have Huawei Cloud accounts, what runs in them today? Would you accept a fresh set of Landing Zone accounts, with existing workloads migrated or redeployed into them?",
   "Greenfield accounts give the cleanest governance baseline. If some accounts must be adopted as-is, list them and what they host.",
   ["01_Foundation.Settings", "01_Foundation.CoreAccounts"],
   "Greenfield: a new organization with fresh core + workload accounts.")
_q("core", "Organization & Accounts",
   "Which applications or workloads move to Huawei Cloud in the first wave? For each, which environments do you run (dev / test / UAT / production), and should each environment be isolated in its own account?",
   "Appendix A (Accounts & Environments) has a table for this if you prefer rows over prose. Per-environment accounts are our recommended isolation boundary.",
   ["01_Foundation.WorkloadAccounts", "01_Foundation.OrganizationalUnits"],
   "One account per workload per environment tier.")
_q("core", "Organization & Accounts",
   "If you built a landing zone or account hierarchy on another cloud (AWS OUs, Azure management groups, GCP folders), describe it. What worked well and what would you change?",
   "A diagram or export is ideal. We mirror proven structure where it fits Huawei Organizations.",
   ["01_Foundation.OrganizationalUnits"],
   example="AWS Control Tower: Security / Infrastructure / Workloads OUs. Worked well; too many SCP exceptions crept in.")
_q("core", "Organization & Accounts",
   "Which Huawei Cloud region is your primary deployment region, and do you have secondary-region or disaster-recovery ambitions? Any data-residency constraints on where data may live?",
   "e.g. ap-southeast-3 (Singapore) primary. Residency constraints become an org-wide region guardrail.",
   ["Global.Settings.home_region", "03_Identity.RegisteredRegions",
    "04_Perimeter.SCPs.AllowedRegions"],
   "Single region ap-southeast-3, region-lock guardrail staged (not enforced).")
_q("core", "Organization & Accounts",
   "What email convention should account mailboxes follow? Each Huawei account needs a globally unique email address - do you have a pattern (e.g. cloud-<account>@yourco.com)?",
   "Plus-addressing (team+lz-prod@yourco.com) works if your mail system supports it.",
   ["01_Foundation.CoreAccounts.Email", "01_Foundation.WorkloadAccounts.Email"],
   "cloud-<account>@<your-domain> pattern proposed in the draft.",
   example="cloud-<account>@acme.com, e.g. cloud-lz-prod@acme.com (plus-addressing not supported).")
_q("core", "Organization & Accounts",
   "Do you have a cloud resource naming convention (VPCs, subnets, gateways, buckets, vaults, log groups, ...)? State the pattern and an example, e.g. <org>-<region>-<env>-<service>-<nn> -> acme-sg-prd-vpc-01.",
   "Every named resource in the draft design is generated from this pattern, so one answer names the whole estate consistently. Note any hard limits (bucket names are globally unique and lowercase; some services cap length).",
   ["05_Network.HubVPCs.VPCName", "05_Network.HubSubnets.Name",
    "05_Network.SpokeVPCs.VPCName", "05_Network.EIPs.Name",
    "05_Network.NATGateways.Name", "06_Observability.AuditSettings",
    "08_DNS.ResolverEndpoints.Name"],
   "Huawei best-practice naming (<org>-<region>-<env>-<service>-<nn>); all resource names inferred from it in the draft spec, flagged for review.",
   example="<org>-<region>-<env>-<service>-<nn>, e.g. acme-sg-prd-vpc-01; buckets lowercase with acme- prefix.")
_q("deep", "Organization & Accounts",
   "How do you expect the account estate to grow over the next 12-24 months - new workloads, business units, or countries? Who approves creating a new account?",
   "Growth expectations size the OU structure and the account-vending process.",
   ["01_Foundation.OrganizationalUnits", "01_Foundation.WorkloadAccounts"])

# ── Identity & Access ───────────────────────────────────────────────────────
_q("core", "Identity & Access",
   "Which teams will work in the cloud (platform, application, security, network, database, finance, ...) and what is each team responsible for?",
   "Appendix C (Teams & Access) has a table for this. Teams map to Identity Center groups and permission sets.",
   ["03_Identity.Groups"],
   example="Platform team (LZ + network), AppDev (portal), SecOps (SOC + audit), DBA; finance needs read-only cost views.")
_q("core", "Identity & Access",
   "How should people sign in to Huawei Cloud: federated through your corporate identity provider (Entra ID / AD FS / Okta / other), or as natively managed Identity Center users? Which IdP product and version do you run?",
   "Federation keeps joiner-mover-leaver in your IdP. Native IC users need no IdP integration but are managed separately.",
   ["03_Identity.Settings", "03_Identity.Users"],
   "Native Identity Center users until IdP federation is confirmed.")
_q("core", "Identity & Access",
   "Does your identity provider support SAML 2.0 for sign-in and SCIM for automatic user/group provisioning?",
   "Both supported: full federation with auto-provisioning. SAML only: federation with manual user sync.",
   ["03_Identity.Settings"],
   example="Entra ID - SAML 2.0 and SCIM both available.")
_q("core", "Identity & Access",
   "Do third parties (vendors, MSPs, auditors, outsourced developers) need access? Who, to which accounts or applications, at what permission level, and standing or time-boxed?",
   "Third-party access gets its own groups and least-privilege permission sets so it can be revoked cleanly.",
   ["03_Identity.PermissionSets", "03_Identity.AccountAssignments"],
   "No third-party access provisioned.")
_q("core", "Identity & Access",
   "What password, MFA and session policies must the cloud enforce? (minimum length, expiry, MFA mandatory for whom, console session timeout, lockout rules)",
   "If you have no written policy, we apply the LZ baseline: 12+ character passwords, 90-day expiry, MFA required, 8-hour sessions, 60-minute console timeout.",
   ["03_Identity.Settings.ic_min_password_length", "03_Identity.Settings.ic_mfa_required"],
   "LZ baseline hardening policy.")
_q("deep", "Identity & Access",
   "Describe your CI/CD and automation landscape: which pipeline tools (GitHub Actions, GitLab, Jenkins, ...), where they run, and how automation should authenticate to Huawei Cloud (short-lived federated credentials vs long-lived access keys).",
   "We strongly prefer OIDC/assume-role federation for pipelines; long-lived AK/SK is a last resort with rotation requirements.",
   ["03_Identity.PermissionSets", "03_Identity.AppPermissionSets"],
   example="GitHub Actions (cloud-hosted); OIDC federation preferred, no static keys in pipelines.")
_q("deep", "Identity & Access",
   "How is emergency (break-glass) access handled today? Who holds root/master credentials, are they MFA-protected, and how is their use audited?",
   "Informs custody of the management-account root user and the break-glass runbook.",
   ["03_Identity.Settings"],
   example="Root credentials sealed with IT security, MFA on; every use requires a ticket and is reviewed.")
_q("deep", "Identity & Access",
   "How granular should permissions be? Are standard tiers (administrator / power user / read-only) per account enough, or do specific teams need permissions scoped to a single application's resources within a shared account?",
   "App-scoped permission sets use enterprise-project scoping - more setup, tighter blast radius. Also note preferred session duration and a portal alias (the sign-in URL name) if you have one.",
   ["03_Identity.PermissionSets", "03_Identity.AppPermissionSets",
    "03_Identity.AccountAssignments", "03_Identity.Settings.identity_center_alias"],
   "Three standard tiers per account; no app-scoped sets.")

# ── Network ─────────────────────────────────────────────────────────────────
_q("core", "Network",
   "Describe your current network: data centres, offices, existing cloud networks, and how they interconnect. Attach a topology diagram if one exists.",
   "Even a rough sketch helps. This anchors the hub-and-spoke design and interconnect plan.",
   ["05_Network.HubVPCs"],
   example="Two DCs (SG + KL) linked by MPLS; AWS VPCs reach on-prem via Transit Gateway; offices on SD-WAN.")
_q("core", "Network",
   "Which connectivity do you need between Huawei Cloud and your on-premises sites or other clouds? Private line (Direct Connect), site-to-site VPN, or both - and does the link need high availability?",
   "List each site/cloud that must reach Huawei Cloud. VPN specifics (devices, IPs, routing) come in the deep-dive section.",
   ["10_VPN.Gateways", "10_VPN.Connections", "10_VPN.Settings"],
   "No hybrid connectivity provisioned.")
_q("core", "Network",
   "How should internet access work? Which applications need outbound internet, which must be reachable from the internet, and is centralized egress through a hub NAT + firewall acceptable?",
   "Centralized egress is the LZ default: all spoke traffic exits through the inspected hub. Exceptions need a stated reason.",
   ["05_Network.NATGateways", "05_Network.SNATRules", "05_Network.Settings.snat_vpc_attachment"],
   "Centralized egress via hub NAT behind the cloud firewall.")
_q("core", "Network",
   "The design connects VPCs hub-and-spoke through an Enterprise Router, which charges per attachment and per GB of traffic. Any objection, or constraints on inter-VPC bandwidth or cost?",
   "The alternative (VPC peering mesh) does not scale with accounts and bypasses central inspection.",
   ["05_Network.EnterpriseRouter", "05_Network.Settings"],
   "Enterprise Router hub-and-spoke accepted.")
_q("core", "Network",
   "IP addressing: can you allocate Huawei Cloud a dedicated private supernet (e.g. a /16 or larger) that overlaps nothing on-premises or in other clouds? Should your team or Huawei plan the per-VPC subnets within it?",
   "Appendix B (IP Plan) has a table for known allocations and ranges to avoid. We recommend one supernet, centrally carved.",
   ["05_Network.Settings.spoke_private_supernet", "05_Network.HubVPCs",
    "05_Network.HubSubnets", "05_Network.SpokeVPCs", "05_Network.SpokeSubnets"],
   "Huawei plans subnets within a customer-provided supernet.",
   example="10.20.0.0/16 reserved for Huawei Cloud (on-prem uses 10.0-10.19); Huawei plans the subnets.")
_q("core", "Network",
   "How many VPCs do you expect per account - one per account, one per environment, or several per application? Any reason to deviate from one VPC per workload account?",
   "One VPC per workload account is the LZ default; the account boundary already isolates environments.",
   ["05_Network.SpokeVPCs"],
   "One VPC per workload account.")
_q("core", "Network",
   "Which applications must communicate with each other, and which must be isolated? If you have a communication matrix (source, destination, port, protocol), attach it.",
   "This drives firewall policy and security-group rules. No matrix: state the rule of thumb, e.g. 'prod isolated from non-prod; shared services reachable by all'.",
   ["09_CFW.ACLRules", "09_CFW.AddressGroups", "09_CFW.ServiceGroups",
    "11_SGACL.SecurityGroups", "11_SGACL.SGRules"],
   "Default-deny between applications; shared services reachable by all.")
_q("core", "Network",
   "Which tags must every cloud resource carry (e.g. project, owner, environment, cost-centre), and should resource creation be blocked when mandatory tags are missing?",
   "Tag keys/values feed cost allocation and the tag-enforcement guardrail. Blocking untagged creation is strict but keeps the estate clean from day one.",
   ["Global.MasterDefaultTags", "01_Foundation.TagPolicies",
    "04_Perimeter.PredefinedTags", "04_Perimeter.SCPs.MandatoryTags",
    "01_Foundation.Settings.enforce_tag_keys_scp"],
   "project/owner/env/bu tagged by the platform; enforcement staged, not blocking.")
_q("deep", "Network",
   "For each VPN site: what VPN device (vendor/model), does it have a static public IP, BGP or static routing (and your on-prem AS number), which on-prem subnets must be reachable, expected throughput, and do you want dual tunnels for HA?",
   "One answer block per site is fine. BGP with dual tunnels is our recommended HA pattern.",
   ["10_VPN.Gateways", "10_VPN.CustomerGateways", "10_VPN.Connections",
    "05_Network.EnterpriseRouter.er_asn"],
   example="HQ: Fortinet 200F, static IP, BGP (AS 65010), reach 10.0.0.0/12, ~200 Mbit/s, dual tunnels.")
_q("deep", "Network",
   "If you plan multiple regions: which traffic flows between them (replication, DR failover, user traffic) and with what bandwidth expectations?",
   "Skip if single-region.",
   [],
   example="Skip - single region for now; DR ambition revisited next year.")
_q("deep", "Network",
   "Within each VPC, how should subnets be divided - by tier (web / app / db), by AZ, by size? Any sizing rules you follow today?",
   "Default: an app subnet and a db subnet per VPC, plus small platform subnets the design adds automatically.",
   ["05_Network.SpokeSubnets", "05_Network.HubSubnets"],
   "app + db subnets per spoke VPC.")
_q("deep", "Network",
   "Do you run or plan Kubernetes/containers (CCE)? If yes: which workloads, and any preference on container networking? Container clusters consume IP space quickly - factor this into the IP plan.",
   "Skip if no containers planned.",
   ["05_Network.SpokeVPCs"],
   example="CCE planned for the portal re-platform next year; reserve a /20 per cluster in the IP plan.")
_q("deep", "Network",
   "DNS: which internal domains do you use, where are the authoritative servers, and must cloud workloads resolve on-prem names (or on-prem resolve cloud names)? Any public domains you want hosted on Huawei Cloud DNS?",
   "e.g. 'corp.internal on two on-prem DCs; cloud must resolve it; keep public zones at our registrar'. This shapes private zones, resolver endpoints and forwarding rules. Note whether DNS queries should be logged for security analytics.",
   ["08_DNS.Settings", "08_DNS.PublicZones", "08_DNS.PrivateZones",
    "08_DNS.RecordSets", "08_DNS.ResolverEndpoints", "08_DNS.ResolverRules",
    "08_DNS.AccessLogs", "05_Network.Settings.subnet_dns_servers"],
   "Private zone + hybrid resolver; public DNS stays at the current registrar.")
_q("deep", "Network",
   "For internet-facing applications: which are published inbound, where should TLS terminate (load balancer / WAF / the app), and do you bring your own certificates?",
   "Each published app becomes a load-balancer/DNAT entry behind the firewall, optionally fronted by WAF.",
   ["05_Network.DNATRules", "05_Network.ELBs"],
   "No inbound publishing provisioned.")
_q("deep", "Network",
   "Expected internet bandwidth: rough outbound Mbit/s at peak, and how many public IPs you expect to need. Prefer paying by bandwidth (flat) or by traffic (per GB)?",
   "Sizes the egress/ingress EIPs. Default: one 100 Mbit/s bandwidth-billed egress EIP, grown on demand.",
   ["05_Network.EIPs"],
   "1x 100 Mbit/s egress EIP, bandwidth-billed.")
_q("deep", "Network",
   "Do you need VPC flow logs (connection records for every VPC) for troubleshooting or security analytics, and how long should they be kept?",
   "Flow logs add LTS log volume/cost. Default: enabled, 90-day hot retention.",
   ["05_Network.Settings.enable_vpc_flow_logs", "05_Network.Settings.flow_log_retention_days"],
   "Flow logs on, 90-day retention.")

# ── Security ────────────────────────────────────────────────────────────────
_q("core", "Security",
   "Which security products do you use today (perimeter firewalls, EDR, SIEM, vulnerability management, PAM), and which must carry over or integrate with the cloud?",
   "e.g. an existing PAM/bastion that must reach cloud servers, or an IPS policy to replicate.",
   ["09_CFW.Settings"],
   example="Palo Alto perimeter, CrowdStrike EDR, Splunk SIEM - EDR and SIEM must extend to cloud servers.")
_q("core", "Security",
   "Cloud Firewall expectations: should intrusion prevention observe first or block immediately? Do you want antivirus scanning at the perimeter? Should the firewall notify on attack detections and high traffic - and to whom? Billing: pay-per-use or monthly subscription?",
   "Default rollout: IPS in observe mode first, then block after tuning; attack + high-traffic alerts to the ops mailbox; pay-per-use until sizing is proven.",
   ["05_Network.CloudFirewall", "09_CFW.Settings.enable_attack_alarm",
    "09_CFW.Settings.enable_traffic_alarm", "09_CFW.Settings.alarm_topic_name",
    "09_CFW.Settings.enable_anti_virus"],
   "IPS observe mode, alarms on to the ops topic, pay-per-use.")
_q("core", "Security",
   "Do you need a cloud security-operations workspace (SecMaster) for centralized alerts and incident handling? Host security (HSS) on servers? Database security (DBSS)? Who consumes the alerts day-to-day?",
   "SecMaster carries a monthly cost per workspace; HSS/DBSS are per-instance quotas.",
   ["07_Security.Settings", "07_Security.SecMasterModules", "07_Security.AlertRules"],
   "SecMaster in the security account; HSS/DBSS deferred.")
_q("core", "Security",
   "Which actions should be hard-blocked organization-wide, for everyone including admins? Examples: leaving the organization, disabling audit logging, making storage public, creating resources outside approved regions.",
   "These become service control policies (guardrails). We stage them in dry-run first, then enforce.",
   ["04_Perimeter.SCPs"],
   "The 8 LZ baseline guardrails, staged (created, not enforced).")
_q("deep", "Security",
   "Web Application Firewall: which domains/applications need WAF protection, expected bandwidth and requests-per-second, and how are TLS certificates managed today?",
   "Skip if no internet-facing web apps. A dedicated WAF instance lives in the hub DMZ in front of the ingress load balancer.",
   ["07_Security.WAF", "07_Security.WAFDomains"],
   "WAF off until an internet-facing app needs it.")
_q("deep", "Security",
   "DDoS: do any public endpoints need tuned traffic-cleaning thresholds or alerting when mitigation kicks in? Any history of DDoS incidents?",
   "Basic Anti-DDoS is free per public IP; we tune thresholds and alarms per EIP.",
   ["07_Security.AntiDDoS"],
   "Default Anti-DDoS thresholds, no per-EIP tuning.")
_q("deep", "Security",
   "Do you maintain IP or domain block/allow lists (threat feeds, geo-blocking, known-bad IPs) that the cloud firewall should enforce from day one?",
   "Attach the lists or name the feed. These seed firewall blacklists/whitelists and domain groups.",
   ["09_CFW.BlackWhiteLists", "09_CFW.DomainGroups"],
   "No seed lists; firewall starts with the rule baseline only.")
_q("deep", "Security",
   "Encryption and key management: any requirements on who controls encryption keys (customer-managed KMS, BYOK), key rotation, or HSM-backed keys?",
   "The LZ encrypts audit/log storage with dedicated KMS keys by default; BYOK changes key custody.",
   ["06_Observability.AuditSettings.kms_audit_alias"],
   "Platform-managed KMS keys per sensitive bucket.")
_q("deep", "Security",
   "Is a public storage bucket ever legitimate for you (static websites, public downloads)? If yes, how should exceptions be approved and marked?",
   "Default guardrail denies public storage outright; a tag-based exception path can be added with an approval process.",
   ["04_Perimeter.SCPs.ExceptionTagKey"],
   "Public storage denied outright, no exception path.")

# ── Compliance & Audit ──────────────────────────────────────────────────────
_q("core", "Compliance & Audit",
   "Must every action in every account be logged centrally, tamper-proof and searchable? How long must audit logs be retained (hot and archived)?",
   "Common answer: 1 year online + long-term archive. Drives the central audit trail, its bucket and retention tiers.",
   ["06_Observability.AuditSettings"],
   "Org-wide audit trail to the security account, 365-day retention.")
_q("core", "Compliance & Audit",
   "Which regulatory or industry frameworks apply to you (e.g. MAS TRM, PDPA, PCI-DSS, ISO 27001, SOC 2), and do auditors need evidence reports from the cloud platform?",
   "Frameworks map to continuous-compliance rule packs evaluated against all accounts.",
   ["04_Perimeter.ConfigConformancePacks"],
   "Landing Zone best-practice pack only.",
   example="ISO 27001 + PDPA; auditors want a yearly conformance evidence export.")
_q("core", "Compliance & Audit",
   "Do you run continuous compliance monitoring today (AWS Config rules, Azure Policy)? Which account/team should own compliance tooling and its findings on Huawei Cloud?",
   "Typically the security account owns the recorder, aggregator and rule packs.",
   ["04_Perimeter.ConfigSetup"],
   "Config recorder + org aggregator in the security account.")
_q("deep", "Compliance & Audit",
   "Should any accounts be exempt from compliance rule packs (sandboxes, short-lived POC accounts)? What makes an account exempt?",
   "Exemptions reduce noise but create blind spots; name them explicitly.",
   ["04_Perimeter.ConfigConformancePacks.ExcludedAccounts"],
   "No exemptions.")

# ── Operations & Monitoring ─────────────────────────────────────────────────
_q("core", "Operations & Monitoring",
   "How do you monitor infrastructure today: which metrics and logs are collected, with which tools (Zabbix, Prometheus, vendor consoles), and what alarm rules and on-call flow exist?",
   "We map existing practice onto Cloud Eye one-click monitoring bundles + an alerting topic.",
   ["06_Observability.OpsSettings", "06_Observability.OneClickNamespaces"],
   "Cloud Eye baseline bundles on core services, alerts to one ops topic.")
_q("core", "Operations & Monitoring",
   "Should logs and monitoring from all accounts aggregate into one central operations account? Which log types matter most (audit, firewall, DNS, flow, application)?",
   "Central aggregation is the LZ default: member-account logs converge to the ops account and archive to object storage.",
   ["06_Observability.LogAggregation"],
   "All platform logs converge centrally and archive to OBS.")
_q("core", "Operations & Monitoring",
   "Log retention: how long must each class of log stay hot/searchable, and how long in cheap archive storage? Any class that must move to deep archive after a period?",
   "e.g. searchable 90 days, archived 1 year, deep-archived beyond. Drives retention tiers per log class.",
   ["06_Observability.LogAggregation.archive_retention_days",
    "06_Observability.AuditSettings.audit_retention_days"],
   "90-day hot, 365-day archive.")
_q("deep", "Operations & Monitoring",
   "Who should receive platform alerts, and how? List email addresses (or SMS numbers) per team/severity. Note: email subscribers must click a confirmation link before alerts flow.",
   "These become subscriptions on the central alerting topic (also used by firewall and DDoS alarms).",
   ["06_Observability.Subscribers", "06_Observability.OpsSettings.topic_name"],
   example="ops-team@acme.com (all severities); oncall SMS +65 9xxx xxxx (critical only).")
_q("deep", "Operations & Monitoring",
   "Do you integrate with a SIEM (Splunk, Sentinel, QRadar)? Which logs must reach it and by which ingestion method?",
   "Cloud logs can be pulled from the central archive or streamed; the method affects the aggregation design.",
   ["06_Observability.LogAggregation.enable_log_aggregation"],
   "No SIEM integration.")
_q("deep", "Operations & Monitoring",
   "Backup and disaster recovery: what must be backed up (servers, databases, file/object data), with what RPO/RTO and retention? Any long-term archive requirement (e.g. move year-old backups to deep archive)?",
   "The LZ provisions the platform; workload backup policy is designed against these answers and lands with the workload rollout.",
   [],
   "Recorded for the workload phase; no platform default.",
   example="All servers nightly (RPO 24h); databases RPO 15 min; keep 30 days hot + 1 year archive.")

# ── Finance & Cost Management ───────────────────────────────────────────────
_q("core", "Finance & Cost Management",
   "Who manages cloud spend? Is there a financial administrator, and are budgets issued and controlled from the management account?",
   "Shapes billing-account setup and who gets cost visibility.",
   ["02_Finance.Settings"],
   "Central billing from the management account.")
_q("core", "Finance & Cost Management",
   "Must cloud costs be allocated within your organization? Along which dimensions - business unit, department, project, environment, owner?",
   "Dimensions become cost-centre enterprise projects + the tag plan from the Network section.",
   ["02_Finance.CostCenters"],
   "Cost centres per business unit, prod/dev typed.")
_q("core", "Finance & Cost Management",
   "How should resources be grouped INSIDE accounts using Enterprise Projects - Huawei's in-account grouping construct (closest analogue: Azure resource groups, or AWS tag-based resource groups)? By application, environment, department, cost centre - or do you have no preference?",
   "Enterprise projects scope both cost reporting and permissions (an admin can be limited to one EP), so the grouping outlives billing. No preference: the design derives an EP layout from your workload and cost-allocation answers.",
   ["02_Finance.CostCenters", "03_Identity.AppPermissionSets"],
   "One EP per application per environment tier, derived from the workload list; flagged for review.",
   example="Group by app + environment, e.g. portal-prd-ep / portal-uat-ep - mirrors our Azure resource-group layout.")
_q("deep", "Finance & Cost Management",
   "How should shared costs be handled - platform accounts (network hub, security, logging) and shared resources used by many departments? Showback, chargeback, or absorbed centrally?",
   "Common pattern: platform costs absorbed centrally or split pro-rata by consumption.",
   ["02_Finance.CostCenters"],
   "Platform costs absorbed centrally.")
_q("deep", "Finance & Cost Management",
   "Do you have an existing Huawei Cloud commercial arrangement (enterprise agreement, partner/reseller, committed spend), and do you want budget alerts at defined thresholds sent to finance?",
   "Commercial structure can constrain how accounts attach to billing; budget alerts reuse the ops alerting topic.",
   ["06_Observability.Subscribers"],
   example="Partner/reseller agreement via <partner>; alert finance@acme.com at 80% of monthly budget.")

# ── Appendices (optional structured tables) ─────────────────────────────────
APPENDICES = [
    {
        "ref": "A", "name": "Appendix A - Accounts",
        "title": "Accounts & Environments (optional)",
        "note": "One row per workload per environment. Fill what you know; leave the rest blank.",
        "columns": ["Application / Workload", "Environment", "Proposed account name",
                    "Grouping / OU", "Owner email", "Notes"],
        "example": ["Customer portal", "prod", "app-portal-prod", "Workloads/Prod",
                    "cloud-portal-prod@example.com", "Internet-facing"],
        "wiring": ["01_Foundation.WorkloadAccounts", "01_Foundation.OrganizationalUnits"],
    },
    {
        "ref": "B", "name": "Appendix B - IP Plan",
        "title": "IP Plan (optional)",
        "note": "Known allocations and ranges to avoid. Scope: supernet offered to Huawei Cloud, an existing on-prem/other-cloud range to avoid, or a specific VPC you want pinned.",
        "columns": ["Scope", "CIDR", "Purpose / VPC name", "Account", "Notes"],
        "example": ["supernet", "10.20.0.0/16", "Huawei Cloud allocation", "-",
                    "Nothing else uses this block"],
        "wiring": ["05_Network.Settings.spoke_private_supernet", "05_Network.HubVPCs",
                   "05_Network.HubSubnets", "05_Network.SpokeVPCs", "05_Network.SpokeSubnets"],
    },
    {
        "ref": "C", "name": "Appendix C - Teams",
        "title": "Teams & Access (optional)",
        "note": "One row per team. Access level: admin / power user / read-only / custom (describe in Notes).",
        "columns": ["Team", "Responsibilities", "Members (name + email)",
                    "Access level", "Accounts or applications in scope", "Notes"],
        "example": ["Platform engineering", "Landing zone, network, shared services",
                    "Alice Tan <alice@example.com>", "admin", "all accounts", ""],
        "wiring": ["03_Identity.Groups", "03_Identity.Users",
                   "03_Identity.PermissionSets", "03_Identity.AccountAssignments"],
    },
]


# ── Coverage check ──────────────────────────────────────────────────────────

def coverage_check():
    """(missing, unknown): schema tables no question reaches / wiring typos."""
    from lz_spec import schema
    valid = set()
    for sh in schema.SHEETS:
        if sh.name in schema.INFO_SHEETS or sh.name == "_meta":
            continue
        for t in sh.tables:
            valid.add(f"{sh.name}.{t.name}")

    covered, unknown = set(), []
    for item in Q + APPENDICES:
        for w in item.get("wiring", []):
            table = ".".join(w.split(".")[:2])
            if table not in valid:
                unknown.append(w)
            covered.add(table)

    missing = sorted(t for t in valid if t not in covered and t not in ENGINEER_ONLY)
    stale = sorted(t for t in ENGINEER_ONLY if t not in valid)
    return missing, sorted(set(unknown)) + [f"stale exemption: {s}" for s in stale]


# ── Workbook emit ───────────────────────────────────────────────────────────

def _refs():
    """Assign C1../D1.. refs in catalogue order."""
    c = d = 0
    for q in Q:
        if q["tier"] == "core":
            c += 1
            q["ref"] = f"C{c}"
        else:
            d += 1
            q["ref"] = f"D{d}"


def write_workbook(out: Path):
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    # Theme aligned with lz_spec/gen_template.py (the LLD spec workbook)
    DARK = "1F4E79"     # title band (= template TITLE_FILL)
    HDR = "DDEBF7"      # header row (= template HEADER_FILL)
    BAND = "C6E0B4"     # category band (= template section-band green)
    FILL_IN = "FFFCE5"  # response cells (= template VALUE_FILL)
    thin = Side(style="thin", color="B4B4B4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(wrap_text=True, vertical="top")

    def _title(ws, text, ncols):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        c = ws.cell(row=1, column=1, value=text)
        c.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=DARK)
        c.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[1].height = 26

    def _headers(ws, row, headers):
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=i, value=h)
            c.font = Font(name="Calibri", bold=True, size=10)
            c.fill = PatternFill("solid", fgColor=HDR)
            c.alignment = wrap
            c.border = border
        ws.row_dimensions[row].height = 20

    def _est_height(texts_widths, minimum=30):
        lines = 1
        for text, width in texts_widths:
            lines = max(lines, -(-len(text) // width) + text.count("\n"))
        return max(minimum, 14 * lines + 6)

    wb = openpyxl.Workbook()

    # Intro ------------------------------------------------------------------
    ws = wb.active
    ws.title = "Start Here"
    ws.column_dimensions["A"].width = 110
    ws.sheet_view.showGridLines = False
    _title(ws, "Huawei Cloud Landing Zone - Assessment Questionnaire", 1)
    intro = [
        ("", None),
        ("Purpose", "h"),
        ("This questionnaire collects the information needed to design your Huawei Cloud "
         "Landing Zone: the governed multi-account foundation (organization, identity, "
         "network, security, audit and cost management) your workloads will land on. "
         "Your answers drive a draft design and configuration; anything unanswered "
         "becomes an explicit decision list we walk through together.", None),
        ("", None),
        ("How to answer", "h"),
        ("- Answer in free text, in as much or as little detail as you have. Attach "
         "diagrams, inventories or policies wherever they exist - a document beats "
         "retyping.", None),
        ("- 'Unknown' and 'no requirement' are valid answers. Please answer every "
         "question on the Core Questions sheet; answer Deep-Dive questions only where "
         "they apply to you.", None),
        ("- Anything left blank on the Deep-Dive sheet is designed to a documented "
         "Huawei best-practice default, which you review before anything is built.", None),
        ("- Appendices A-C are optional tables for facts that fit rows better than "
         "prose (accounts, IP ranges, teams). Fill what you know.", None),
        ("", None),
        ("What happens next", "h"),
        ("1. We convert your answers into a draft Landing Zone configuration.", None),
        ("2. Open points and defaults come back to you as a short decision list.", None),
        ("3. A workshop resolves the decision list and finalizes the design (LLD).", None),
        ("4. The landing zone is built from that configuration - no console clicking.", None),
        ("", None),
        (f"Version {QUESTIONNAIRE_VERSION} - generated {datetime.date.today():%Y-%m-%d}"
         f" - schema {_schema_version()}", "s"),
    ]
    r = 2
    for text, kind in intro:
        c = ws.cell(row=r, column=1, value=text or None)
        if kind == "h":
            c.font = Font(bold=True, size=11, color=DARK)
        elif kind == "s":
            c.font = Font(italic=True, size=9, color="595959")
        c.alignment = wrap
        if text and kind is None:
            ws.row_dimensions[r].height = _est_height([(text, 105)], minimum=15)
        r += 1

    # Survey sheets ----------------------------------------------------------
    widths = {"A": 7, "B": 46, "C": 40, "D": 34, "E": 50}
    for tier, sheet_name, subtitle in [
        ("core", "Core Questions", "Core Questions - please answer all"),
        ("deep", "Deep-Dive Questions", "Deep-Dive Questions - answer where applicable; blank = best-practice default"),
    ]:
        ws = wb.create_sheet(sheet_name)
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        _title(ws, f"Huawei Cloud Landing Zone Assessment - {subtitle}", 4)
        _headers(ws, 2, ["No.", "Question", "Guidance", "Example Response", "Customer Response"])
        ws.freeze_panes = "A3"
        r = 3
        for cat in CATEGORIES:
            qs = [q for q in Q if q["tier"] == tier and q["category"] == cat]
            if not qs:
                continue
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
            c = ws.cell(row=r, column=1, value=cat)
            c.font = Font(name="Calibri", bold=True, color="375623")
            c.fill = PatternFill("solid", fgColor=BAND)
            c.alignment = Alignment(vertical="center", indent=1)
            ws.row_dimensions[r].height = 18
            r += 1
            for q in qs:
                ex = q["example"] or q["default"]
                cells = [q["ref"], q["question"], q["guidance"], ex or None, None]
                for i, v in enumerate(cells, 1):
                    c = ws.cell(row=r, column=i, value=v)
                    c.alignment = wrap
                    c.border = border
                    if i == 4:
                        c.font = Font(italic=True, size=10, color="595959")
                    if i == 5:
                        c.fill = PatternFill("solid", fgColor=FILL_IN)
                ws.row_dimensions[r].height = _est_height(
                    [(q["question"], 42), (q["guidance"], 36), (ex, 30)])
                r += 1

    # Appendix sheets --------------------------------------------------------
    for ap in APPENDICES:
        ws = wb.create_sheet(ap["name"])
        n = len(ap["columns"])
        for i in range(1, n + 1):
            ws.column_dimensions[get_column_letter(i)].width = 30
        _title(ws, ap["title"], n)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n)
        c = ws.cell(row=2, column=1, value=ap["note"])
        c.font = Font(italic=True, size=9, color="595959")
        c.alignment = wrap
        ws.row_dimensions[2].height = 28
        _headers(ws, 3, ap["columns"])
        ws.freeze_panes = "A4"
        for i, v in enumerate(ap["example"], 1):
            c = ws.cell(row=4, column=i, value=(f"(example) {v}" if i == 1 else v) or None)
            c.font = Font(italic=True, color="595959")
            c.alignment = wrap
            c.border = border
        for r in range(5, 25):
            for i in range(1, n + 1):
                c = ws.cell(row=r, column=i)
                c.alignment = wrap
                c.border = border

    # Hidden wiring sheet ----------------------------------------------------
    ws = wb.create_sheet("_wiring")
    ws.append(["Ref", "Tier", "Category", "Targets", "DefaultIfSilent"])
    for q in Q:
        ws.append([q["ref"], q["tier"], q["category"],
                   "; ".join(q["wiring"]), q["default"]])
    for ap in APPENDICES:
        ws.append([ap["ref"], "appendix", ap["title"],
                   "; ".join(ap["wiring"]), ""])
    ws.append(["_meta", "meta", "",
               f"questionnaire_version={QUESTIONNAIRE_VERSION}; schema_version={_schema_version()}", ""])
    ws.sheet_state = "hidden"

    wb.save(out)


def _schema_version():
    from lz_spec import schema
    return schema.SCHEMA_VERSION


def main(argv=None):
    ap = argparse.ArgumentParser()
    ap.add_argument("-o", "--out", default=str(DEFAULT_OUT))
    ap.add_argument("--check", action="store_true", help="coverage check only, write nothing")
    args = ap.parse_args(argv)

    _refs()
    missing, unknown = coverage_check()
    for m in missing:
        print(f"ERROR: schema table not covered by any question: {m}", file=sys.stderr)
    for u in unknown:
        print(f"ERROR: wiring target not in schema: {u}", file=sys.stderr)
    if missing or unknown:
        return 1
    core = sum(1 for q in Q if q["tier"] == "core")
    print(f"coverage OK: {core} core + {len(Q) - core} deep questions, "
          f"{len(APPENDICES)} appendices, {len(ENGINEER_ONLY)} exempt tables")
    if args.check:
        return 0
    write_workbook(Path(args.out))
    print(f"wrote {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
