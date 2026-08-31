"""Generate the as-built build-configuration book from tfvars (+ optional states).

Customer-agnostic: every value comes from the envs tree's terraform.tfvars.json
files and, when --states-dir is given, from pulled state (live IDs/EIPs).
Sections degrade gracefully when a feature is unused or state is absent.

Usage:
    py tools/gen_config_book.py --envs-dir envs \
        [--states-dir <dir>] --out book.xlsx [--customer "Example Corp"] [--version 1.0]
"""

import argparse
import json
import sys
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from envtree import (BOX, HDR_FILL, HDR_FONT, SEC_FILL, SEC_FONT, WRAP,
                     env_dirs, tfvars, state, instances)



class Book:
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)

    def sheet(self, name, widths):
        return _Sheet(self.wb.create_sheet(name), widths)


class _Sheet:
    def __init__(self, ws, widths):
        self.ws = ws
        for i, w in enumerate(widths, 1):
            self.ws.column_dimensions[get_column_letter(i)].width = w
        self.r = 1

    def title(self, text):
        self.ws.cell(row=self.r, column=1, value=text).font = Font(bold=True, size=14)
        self.r += 2

    def section(self, text):
        c = self.ws.cell(row=self.r, column=1, value=text)
        c.font, c.fill = SEC_FONT, SEC_FILL
        self.r += 1

    def table(self, headers, rows):
        for i, h in enumerate(headers, 1):
            c = self.ws.cell(row=self.r, column=i, value=h)
            c.font, c.fill, c.border = HDR_FONT, HDR_FILL, BOX
        self.r += 1
        for row in rows:
            for i, v in enumerate(row, 1):
                c = self.ws.cell(row=self.r, column=i, value=v)
                c.border, c.alignment = BOX, WRAP
            self.r += 1
        self.r += 1

    def kv(self, pairs):
        self.table(["Setting", "Value"], [(k, "" if v is None else str(v)) for k, v in pairs])


def _j(v, n=200):
    return json.dumps(v)[:n] if v is not None else ""


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--envs-dir", required=True)
    ap.add_argument("--states-dir")
    ap.add_argument("--out", required=True)
    ap.add_argument("--customer", default="")
    ap.add_argument("--version", default="1.0")
    args = ap.parse_args()

    envs_dir = Path(args.envs_dir)
    states_dir = Path(args.states_dir) if args.states_dir else None
    tv = lambda env: tfvars(envs_dir, env)
    st = lambda env: state(states_dir, env) if states_dir else {}
    envs = [d.name for d in env_dirs(envs_dir)]

    book = Book()
    f = tv("01-foundation")
    st01 = st("01-foundation")

    # ---- Overview ----
    s = book.sheet("Overview", [34, 70])
    label = f"{args.customer} Landing Zone - Build Configuration Book" if args.customer \
        else "Landing Zone - Build Configuration Book"
    s.title(label)
    org = instances(st01, "huaweicloud_organizations_organization")
    s.kv(([("Document version", args.version)] if args.version else []) + [
        ("Generated", date.today().isoformat()),
        ("Source of record", "Terraform configuration and state of each environment"),
        ("Region", f.get("home_region", "")),
        ("Organization ID", org[0][2].get("id") if org else "-"),
        ("Identity Center alias", f.get("identity_center_alias", "-")),
        ("Cross-account agency", f.get("cross_account_agency_name", "-")),
        ("Default tags", _j(f.get("default_tags"))),
        ("Apply order", ", ".join(envs)),
    ])

    # ---- 01 Organization ----
    s = book.sheet("01 Organization", [26, 40, 14, 36, 34])
    s.title("Organization structure (envs/01-foundation)")
    core = set((f.get("core_accounts") or {}))
    accts = instances(st01, "huaweicloud_organizations_account")
    if accts:
        s.section("Member accounts (from state)")
        s.table(["Account", "Email", "Type", "Account ID", "Description"],
                [(a.get("name"), a.get("email", ""), "core" if a.get("name") in core else "workload",
                  a.get("id", ""), a.get("description", ""))
                 for _, _, a in sorted(accts, key=lambda x: str(x[2].get("name")))])
    else:
        s.section("Member accounts (from configuration; no state provided)")
        rows = [(n, v.get("email", ""), "core", "-", v.get("description", ""))
                for n, v in (f.get("core_accounts") or {}).items()]
        rows += [(n, v.get("email", ""), "workload", "-", v.get("description", ""))
                 for n, v in (f.get("workload_accounts") or {}).items()]
        s.table(["Account", "Email", "Type", "Account ID", "Description"], sorted(rows))
    s.section("Organizational units")
    s.table(["OU", "Parent"], [(k, (v.get("parent") or "root") if isinstance(v, dict) else "root")
                               for k, v in (f.get("organizational_units") or {}).items()])
    s.section("Trusted services and delegated administrators")
    da = f.get("delegated_administrators") or {}
    s.table(["Service", "Delegated administrator"],
            [(svc, da.get(svc, "-")) for svc in (f.get("trusted_services") or [])])
    s.section("Organization tag policies")
    s.table(["Policy", "Description"],
            [(p.get("name"), p.get("description", "")) for p in (f.get("tag_policies") or [])])

    # ---- 02 Finance ----
    s = book.sheet("02 Finance", [26, 80])
    s.title("Cost centers / enterprise projects (envs/02-finance)")
    cc = tv("02-finance").get("cost_centers_by_account") or {}
    rows = []
    for acct, val in cc.items():
        if isinstance(val, list):
            rows.append((acct, ", ".join((v.get("name", str(v)) if isinstance(v, dict) else str(v)) for v in val)))
        else:
            rows.append((acct, _j(val)))
    s.table(["Account", "Cost-center enterprise projects"], rows)

    # ---- 03 Identity ----
    s = book.sheet("03 Identity", [30, 90])
    i3 = tv("03-identity")
    s.title("Identity and access (envs/03-identity)")
    s.section("Identity Center")
    s.kv([("Session duration", i3.get("session_duration")),
          ("MFA management", _j(i3.get("ic_mfa_management"))),
          ("IC password policy", _j(i3.get("ic_password_policy"), 400)),
          ("Registered regions", ", ".join(i3.get("registered_regions") or []))])
    s.section("Permission sets")
    s.table(["Permission set", "Definition"],
            [(k, _j(v, 500)) for k, v in (i3.get("permission_sets") or {}).items()])
    s.section("Per-account IAM baseline")
    s.kv([("Login policy", _j(i3.get("iam_login_policy"), 400)),
          ("Password policy", _j(i3.get("iam_password_policy"), 400))])

    # ---- 04 Governance ----
    s = book.sheet("04 Governance", [34, 12, 90])
    p4 = tv("04-perimeter")
    s.title("Guardrails, tags and Config (envs/04-perimeter)")
    s.section("SCP guardrails")
    s.table(["Guardrail", "Enforced", "Settings"],
            [(k, str(v.get("enforce", False)),
              _j({x: y for x, y in v.items() if x != "enforce"}, 400))
             for k, v in (p4.get("scps") or {}).items()])
    s.section("Predefined tags (all accounts)")
    s.table(["Tag key", "Allowed values"],
            [(t.get("key"), ", ".join(t.get("values") or []) or "(free text)")
             for t in (p4.get("predefined_tags") or [])])
    s.section("Config (RMS)")
    s.kv([("Config admin account", p4.get("config_admin_account"))]
         + sorted((p4.get("config") or {}).items()))
    s.section("Conformance packages (organization-wide)")
    s.table(["Package", "Enabled", "Template / parameters / exclusions"],
            [(c.get("name"), str(c.get("enabled")),
              _j({k: v for k, v in c.items() if k not in ("name", "enabled")}, 400))
             for c in (p4.get("conformance_packs") or [])])

    # ---- 05 Network ----
    n5 = tv("05-network")
    s = book.sheet("05 Network - Hub", [38, 24, 22, 60])
    hub_acct = n5.get("hub_account", "")
    s.title(f"Hub network (envs/05-network, account {hub_acct})")
    s.section("Core settings")
    keys = ("hub_account", "spoke_private_supernet", "enterprise_project_name",
            "inbound_route_table", "outbound_route_table", "snat_vpc_attachment",
            "enable_vpc_flow_logs", "flow_log_retention_days")
    s.kv([(k, n5.get(k)) for k in keys if k in n5]
         + [("Subnet DNS servers (DHCP)", ", ".join(n5.get("subnet_dns") or []))])
    s.section("Hub VPCs and subnets")
    rows = []
    for vname, v in (n5.get("hub_vpcs") or {}).items():
        rows.append((vname, v.get("cidr"), "VPC", ""))
        for sub in v.get("subnets", []):
            rows.append(("", sub.get("cidr"), "subnet", sub.get("name")))
    s.table(["VPC", "CIDR", "Type", "Subnet name"], rows)
    if n5.get("er_name"):
        s.section("Enterprise Router")
        s.kv([("Name", n5.get("er_name")), ("ASN", n5.get("er_asn")),
              ("Availability zones", ", ".join(n5.get("er_availability_zones") or [])),
              ("Auto-accept shared attachments", n5.get("er_auto_accept_shared_attachments")),
              ("RAM share", n5.get("er_share_name")),
              ("RAM share principals", ", ".join(n5.get("ram_share_principals") or []))])
        s.section("ER route tables")
        s.table(["Route table", "Default route to firewall", "Purpose"],
                [(rt.get("name"),
                  str(rt.get("default_to_cfw", rt.get("name") in (n5.get("cfw_default_route_tables") or []))),
                  rt.get("description", "")) for rt in (n5.get("er_route_tables") or [])])
        s.section("Hub ER attachments")
        s.table(["Attachment", "VPC", "Subnet"],
                [(a.get("name"), a.get("vpc"), a.get("subnet", ""))
                 for a in (n5.get("er_attachments") or [])])
    if n5.get("cfw_name"):
        s.section("Cloud Firewall instance")
        s.kv([(k, n5.get(k)) for k in ("cfw_name", "cfw_flavor", "east_west_firewall_mode",
                                       "inspection_cidr_reservation", "cfw_charging_mode",
                                       "cfw_lts_log_group_name") if k in n5])
    if n5.get("nat_gateways"):
        s.section("NAT gateways")
        s.table(["Name", "Spec", "VPC", "Subnet"],
                [(g.get("name"), g.get("spec", ""), g.get("vpc"), g.get("subnet", ""))
                 for g in n5.get("nat_gateways") or []])
    if n5.get("eips"):
        s.section("Elastic IPs")
        s.table(["Name", "Type", "Billing", "Bandwidth"],
                [(e.get("name"), e.get("type", ""),
                  f"{e.get('billed_by', '')} (pay-per-use)".strip(),
                  f"{e['bandwidth_size']} Mbit/s" if e.get("bandwidth_size") else "")
                 for e in n5.get("eips") or []])
    if n5.get("snat_rules"):
        s.section("SNAT rules")
        s.table(["Source CIDR", "NAT gateway", "EIP"],
                [(r.get("cidr"), r.get("nat_name", ""), r.get("eip", ""))
                 for r in n5.get("snat_rules") or []])

    s = book.sheet("05 Network - Spokes", [30, 20, 26, 26, 12, 44])
    s.title("Spoke networks (envs/05-network)")
    rows = []
    for vname, sp in (n5.get("spokes") or {}).items():
        rows.append((vname, sp.get("account", ""), sp.get("vpc_cidr"), "",
                     str(sp.get("er_attach", True)), _j(sp.get("vpc_tags"), 120)))
        for sub in sp.get("subnets", []):
            rows.append(("", "", sub.get("cidr"), sub.get("name"), "", ""))
    s.table(["Spoke VPC", "Account", "CIDR", "Subnet", "ER attached", "Tags"], rows)

    # ---- 06 Observability ----
    s = book.sheet("06 Observability", [42, 70])
    o6 = tv("06-observability")
    st06 = st("06-observability")
    s.title("Audit, monitoring and log aggregation (envs/06-observability)")
    s.section("Central audit (CTS)")
    s.kv([("Audit bucket", o6.get("audit_bucket_name")),
          ("Audit retention / cold tier",
           f'{o6.get("audit_retention_days")} days delete / COLD at {o6.get("audit_cold_after_days", 0)} days'),
          ("Audit KMS alias", o6.get("kms_audit_alias")),
          ("CTS log group / stream", f'{o6.get("cts_log_group_name")} / {o6.get("cts_log_stream_name")}'),
          ("Minimal trackers (no transfer)", ", ".join(o6.get("cts_no_transfer_accounts") or []) or "-")])
    if o6.get("enable_log_aggregation"):
        s.section("Log aggregation and archive")
        s.kv([("Archive bucket", o6.get("archive_bucket_name")),
              ("Archive retention / cold tier",
               f'{o6.get("archive_retention_days")} days delete / COLD at {o6.get("archive_cold_after_days", 0)} days'),
              ("Archive KMS alias", o6.get("kms_archive_alias")),
              ("Converged hot retention (LTS)", f'{o6.get("converged_retention_days")} days'),
              ("Transfer cadence", f'{o6.get("transfer_period")} {o6.get("transfer_period_unit")}')])
    if st06:
        s.section("Archive transfers (from state)")
        s.table(["Transfer key", "Type"],
                sorted([(str(k), "admin-local direct" if n == "archive_local" else "converged member")
                        for n, k, a in instances(st06, "huaweicloud_lts_transfer")]))
        s.section("Notifications and alarms")
        s.kv([("SMN topics", len(instances(st06, "huaweicloud_smn_topic"))),
              ("SMN subscriptions", len(instances(st06, "huaweicloud_smn_subscription"))),
              ("CES one-click alarm bundles", len(instances(st06, "huaweicloud_ces_one_click_alarm")))])

    # ---- 07 DNS ----
    s = book.sheet("07 DNS", [34, 22, 26, 50])
    d7 = tv("08-network-dns")
    s.title("Hybrid DNS (envs/08-network-dns)")
    if d7.get("resolver_endpoints"):
        s.section("Resolver endpoints")
        s.table(["Endpoint", "Direction", "VPC", "IPs / subnet"],
                [(e.get("name"), e.get("direction"), e.get("vpc", ""),
                  ", ".join(e.get("ips") or []) or e.get("subnet", ""))
                 for e in d7.get("resolver_endpoints") or []])
    if d7.get("resolver_rules"):
        s.section("Forwarding rules")
        s.table(["Rule", "Domain", "Target DNS", "VPCs"],
                [(r.get("name"), r.get("domain_name", ""),
                  ", ".join(str(x) for x in r.get("ip_addresses") or []),
                  ", ".join(r.get("vpcs") or [])) for r in d7.get("resolver_rules") or []])
    if d7.get("access_logs"):
        s.section("Query logging")
        s.table(["Access log", "LTS group", "LTS stream", "VPCs"],
                [(a.get("name"), a.get("lts_group"), a.get("lts_stream"), ", ".join(a.get("vpcs") or []))
                 for a in d7.get("access_logs") or []])
    s.section("Hosted zones")
    s.kv([("Public zones", len(d7.get("public_zones") or []) or "none"),
          ("Private zones", len(d7.get("private_zones") or []) or "none"),
          ("Record sets", len(d7.get("record_sets") or []) or "none")])

    # ---- 08 Firewall ----
    s = book.sheet("08 Firewall", [30, 10, 10, 46, 46, 34, 10])
    c8 = tv("09-network-cfw")
    s.title("Cloud Firewall rule plane (envs/09-network-cfw)")
    if c8.get("address_groups"):
        s.section("Address groups")
        s.table(["Group", "Members"],
                [(g.get("name"), ", ".join(g.get("members") or [])[:900]) for g in c8["address_groups"]])
    if c8.get("domain_groups"):
        s.section("Domain groups")
        s.table(["Group", "Type", "Domains"],
                [(g.get("name"), g.get("type"), ", ".join(g.get("domains") or [])[:900]) for g in c8["domain_groups"]])
    if c8.get("service_groups"):
        s.section("Service groups")
        s.table(["Group", "Members"],
                [(g.get("name"), ", ".join(g.get("members") or [])) for g in c8["service_groups"]])
    if c8.get("acl_rules"):
        s.section("ACL rules (in order; catch-all denies pinned last)")
        s.table(["Rule", "Kind", "Action", "Source", "Destination", "Service", "Status"],
                [(r.get("name"), r.get("kind"), r.get("action"), ", ".join(r.get("source") or [])[:250],
                  ", ".join(r.get("destination") or [])[:250], ", ".join(r.get("service") or [])[:150],
                  r.get("status")) for r in c8["acl_rules"]])

    # ---- 09 VPN ----
    s = book.sheet("09 VPN", [36, 70])
    v9 = tv("05-network")
    st09 = st("05-network")
    s.title("Site-to-cloud VPN (merged into envs/05-network)")
    gws = v9.get("gateways") or []
    if gws:
        live = {a.get("name"): a for _, _, a in instances(st09, "huaweicloud_vpn_gateway")}
        for gw in gws:
            g0 = live.get(gw.get("name"), {})
            s.section(f"VPN gateway {gw.get('name')}")
            s.kv([("Attachment", gw.get("attachment", "er")),
                  ("HA mode", g0.get("ha_mode", gw.get("ha_mode", ""))),
                  ("EIP 1", (g0.get("eip1") or [{}])[0].get("ip_address", "(from state after apply)")),
                  ("EIP 2", (g0.get("eip2") or [{}])[0].get("ip_address", "(from state after apply)")),
                  ("Note", "EIP settings are create-time only; changes replace the gateway with new public IPs")])
        s.section("Customer gateways")
        s.table(["Name", "Peer IP", "Route mode"],
                [(g.get("name"), g.get("ip"), g.get("route_mode"))
                 for g in v9.get("customer_gateways") or []])
        n_conn = len(v9.get("connections") or [])
        n_live = len(instances(st09, "huaweicloud_vpn_connection"))
        s.section("IPsec connections")
        s.kv([("In Terraform", n_conn or n_live or 0),
              ("Note", "Connections not present in Terraform are managed in the console"
               if (n_conn == 0 and n_live == 0) else "")])
    else:
        s.kv([("Status", "No VPN gateways defined")])

    # ---- 10 Security ----
    s = book.sheet("10 Security", [36, 70])
    s10 = tv("07-security")
    st10 = st("07-security")
    s.title("Security services (envs/07-security)")
    applied = bool(st10.get("resources"))
    pairs = [("Applied", "yes" if applied else "no (defined, not yet applied)")]
    if "enable_secmaster" in s10 or any("secmaster" in k for k in s10):
        pairs.append(("SecMaster", "enabled" if s10.get("enable_secmaster") else "configured/disabled"))
    pairs.append(("WAF", "enabled" if s10.get("enable_waf") else "disabled"))
    pairs.append(("Anti-DDoS rows", len(s10.get("antiddos") or [])))
    if s10.get("waf_domains"):
        pairs.append(("WAF protected domains", ", ".join(d.get("domain", "") for d in s10["waf_domains"])))
    s.kv(pairs)

    book.wb.save(args.out)
    print(f"written: {args.out}")
    print("sheets:", book.wb.sheetnames)
    return 0


if __name__ == "__main__":
    sys.exit(main())
