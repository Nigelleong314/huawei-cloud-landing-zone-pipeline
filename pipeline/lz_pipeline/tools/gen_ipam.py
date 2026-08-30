"""Generate the IP-management workbook (block ledger + subnets + hosts) from
the 05-network tfvars of any envs tree. Customer-agnostic.

Usage:
    py tools/gen_ipam.py --envs-dir envs --out ipam.xlsx \
        [--title "Example Landing Zone - IP management"] [--block-prefix 22] \
        [--reserve "10.42.8.0/22=CFW inspection block; never assign"] \
        [--hosts hosts.csv]

--reserve is repeatable for blocks consumed outside Terraform (for example the
firewall ER-mode inspection reservation). --hosts seeds the Hosts sheet from a
CSV (ip,subnet,resource,env,notes); otherwise the sheet ships as an empty
register with headers.
"""

import argparse
import csv
import ipaddress
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.path.insert(0, str(Path(__file__).parent))
from envtree import BOX, HDR_FILL, HDR_FONT, WRAP, tfvars

NOTE_FONT = Font(size=9, italic=True, color="595959")
FILL = {"Free": PatternFill("solid", fgColor="E2EFDA"),
        "Allocated": PatternFill("solid", fgColor="FCE4EC"),
        "Reserved": PatternFill("solid", fgColor="FFF2CC")}


def header(ws, row, cols):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.font, cell.fill, cell.border = HDR_FONT, HDR_FILL, BOX
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def fit(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def collect(n5: dict):
    """(vpcs, subnets) from 05-network tfvars: hub_vpcs + spokes."""
    vpcs, subnets = [], []
    for name, v in (n5.get("hub_vpcs") or {}).items():
        vpcs.append((name, v.get("cidr"), "Hub VPC"))
        for s in v.get("subnets", []):
            subnets.append((name, s.get("name"), s.get("cidr")))
    for name, sp in (n5.get("spokes") or {}).items():
        acct = sp.get("account", "")
        att = "" if sp.get("er_attach", True) else ", detached from the ER"
        vpcs.append((name, sp.get("vpc_cidr"), f"Spoke VPC ({acct}){att}"))
        for s in sp.get("subnets", []):
            subnets.append((name, s.get("name"), s.get("cidr")))
    return vpcs, subnets


def subnet_purpose(name: str) -> str:
    return "ER attachment" if "-att" in (name or "") else ""


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--envs-dir", required=True)
    ap.add_argument("--out", required=True)
    ap.add_argument("--title", default="Landing Zone - IP management")
    ap.add_argument("--block-prefix", type=int, default=22)
    ap.add_argument("--supernet", help="override; default = 05-network spoke_private_supernet")
    ap.add_argument("--reserve", action="append", default=[], metavar="CIDR=note")
    ap.add_argument("--hosts", help="CSV of ip,subnet,resource,env,notes")
    args = ap.parse_args()

    n5 = tfvars(Path(args.envs_dir), "05-network")
    if not n5:
        print("05-network/terraform.tfvars.json not found", file=sys.stderr)
        return 2
    supernet = ipaddress.ip_network(args.supernet or n5["spoke_private_supernet"])
    vpcs, subnets = collect(n5)

    reserved = {}
    for r in args.reserve:
        cidr, _, note = r.partition("=")
        reserved[str(ipaddress.ip_network(cidr.strip()))] = note.strip()

    # map VPC allocations onto the /N carving
    alloc = {}
    for name, cidr, kind in vpcs:
        if not cidr:
            continue
        net = ipaddress.ip_network(cidr)
        for block in supernet.subnets(new_prefix=args.block_prefix):
            if net.overlaps(block):
                alloc[str(block)] = (name, kind)

    wb = openpyxl.Workbook()
    s = wb.active
    s.title = "Summary"
    s["A1"] = args.title
    s["A1"].font = Font(bold=True, size=14)
    n_blocks = 2 ** (args.block_prefix - supernet.prefixlen)
    s["A2"] = (f"Supernet {supernet}, carved into {n_blocks} blocks of /{args.block_prefix}. "
               "Keep the Blocks sheet up to date; every number below recalculates from it.")
    s["A2"].font = NOTE_FONT
    last = n_blocks + 1
    rows = [
        ("Supernet", str(supernet)),
        (f"Total /{args.block_prefix} blocks", f"=COUNTA(Blocks!A2:A{last})"),
        ("Allocated", f'=COUNTIF(Blocks!D2:D{last},"Allocated")'),
        ("Reserved", f'=COUNTIF(Blocks!D2:D{last},"Reserved")'),
        ("Free blocks", f'=COUNTIF(Blocks!D2:D{last},"Free")'),
        ("Next free block", f'=INDEX(Blocks!A2:A{last},MATCH("Free",Blocks!D2:D{last},0))'),
        ("Registered host IPs", "=COUNTA(Hosts!A2:A500)"),
    ]
    for i, (k, v) in enumerate(rows, 4):
        s.cell(row=i, column=1, value=k).font = Font(bold=True)
        s.cell(row=i, column=2, value=v)
    fit(s, [24, 40])

    b = wb.create_sheet("Blocks")
    header(b, 1, [f"Block (/{args.block_prefix})", "First IP", "Last IP", "Status", "Assigned to", "Notes"])
    dv = DataValidation(type="list", formula1='"Allocated,Reserved,Free"', allow_blank=False)
    b.add_data_validation(dv)
    r = 2
    for net in supernet.subnets(new_prefix=args.block_prefix):
        cidr = str(net)
        if cidr in alloc:
            status, owner, note = "Allocated", alloc[cidr][0], alloc[cidr][1]
        elif cidr in reserved:
            status, owner, note = "Reserved", "", reserved[cidr]
        else:
            status, owner, note = "Free", "", ""
        for c, v in enumerate([cidr, str(net[0]), str(net[-1]), status, owner, note], 1):
            cell = b.cell(row=r, column=c, value=v)
            cell.border, cell.alignment = BOX, WRAP
        b.cell(row=r, column=4).fill = FILL[status]
        dv.add(b.cell(row=r, column=4))
        r += 1
    fit(b, [18, 15, 15, 12, 34, 55])

    sn = wb.create_sheet("Subnets")
    header(sn, 1, ["VPC", "Subnet", "CIDR", "Purpose", "Usable IPs", "Host IPs used"])
    for i, (vpc, name, cidr) in enumerate(subnets, 2):
        size = (ipaddress.ip_network(cidr).num_addresses - 5) if cidr else ""
        vals = [vpc, name, cidr, subnet_purpose(name), size, f'=COUNTIF(Hosts!B2:B500,B{i})']
        for c, v in enumerate(vals, 1):
            cell = sn.cell(row=i, column=c, value=v)
            cell.border, cell.alignment = BOX, WRAP
    fit(sn, [30, 38, 18, 22, 11, 14])

    h = wb.create_sheet("Hosts")
    header(h, 1, ["IP", "Subnet", "Resource", "Env", "Notes"])
    if args.hosts:
        with open(args.hosts, newline="", encoding="utf-8") as fh:
            for i, row in enumerate(csv.reader(fh), 2):
                for c, v in enumerate(row[:5], 1):
                    cell = h.cell(row=i, column=c, value=v)
                    cell.border, cell.alignment = BOX, WRAP
    fit(h, [16, 38, 34, 8, 45])

    wb.save(args.out)
    print(f"written: {args.out}  (blocks: {n_blocks}, allocated: {len(alloc)}, "
          f"reserved: {len(reserved)}, subnets: {len(subnets)})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
