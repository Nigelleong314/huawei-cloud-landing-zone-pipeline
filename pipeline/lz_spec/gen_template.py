"""Generate the LDZ spec Excel workbook from schema.SHEETS.

Usage: python gen_template.py <output.xlsx>
"""

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from lz_spec.schema import SHEETS, KV, Table, Sheet


# Styling
TITLE_FILL  = PatternFill("solid", fgColor="1F4E78")
HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
SAMPLE_FILL = PatternFill("solid", fgColor="EFEFEF")
VALUE_FILL  = PatternFill("solid", fgColor="FFFCE5")
SENTINEL_FILL = PatternFill("solid", fgColor="C6E0B4")
NOTE_FILL = PatternFill("solid", fgColor="F2F2F2")

TITLE_FONT  = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
HEADER_FONT = Font(name="Calibri", size=10, bold=True)
SENTINEL_FONT = Font(name="Calibri", size=10, bold=True, color="375623")
NOTE_FONT = Font(name="Calibri", size=9, italic=True, color="595959")
CELL_FONT = Font(name="Calibri", size=10)

THIN = Side(border_style="thin", color="B4B4B4")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

# ── Template v2 guided entry ────────────────────────────────────────────────

# Columns the resolver fills when left blank (grey cells, annotated type row).
# Static enum dropdowns (bool columns get TRUE/FALSE automatically by type).
ENUM_COLS = {
    ("02_Finance", "CostCenters", "EnterpriseProjectType"): "prod,poc",
    ("08_DNS", "ResolverEndpoints", "Direction"): "inbound,outbound",
    ("08_DNS", "RecordSets", "Type"): "A,AAAA,CNAME,MX,TXT,NS,SRV,PTR,CAA",
    ("09_CFW", "ACLRules", "Kind"): "vpc,nat,eip",
    ("09_CFW", "ACLRules", "Action"): "allow,deny",
    ("09_CFW", "ACLRules", "Status"): "enable,disable",
    ("09_CFW", "ACLRules", "Direction"): "inbound,outbound",
    ("09_CFW", "BlackWhiteLists", "ListType"): "black,white",
    ("09_CFW", "BlackWhiteLists", "Direction"): "inbound,outbound",
    ("09_CFW", "BlackWhiteLists", "Protocol"): "tcp,udp,icmp,any",
    ("09_CFW", "BlackWhiteLists", "AddressType"): "ipv4,ipv6",
    ("10_VPN", "Gateways", "Attachment"): "er,vpc",
    ("10_VPN", "Gateways", "ERAssocRouteTable"): "er-inbound,er-outbound,er-hybrid",
    ("10_VPN", "Gateways", "ERPropRouteTable"): "er-inbound,er-outbound,er-hybrid",
    ("10_VPN", "Gateways", "HAMode"): "active-active,active-standby",
    ("10_VPN", "Gateways", "NetworkType"): "public,private",
    ("10_VPN", "CustomerGateways", "RouteMode"): "static,bgp",
    ("10_VPN", "Connections", "VPNType"): "static,bgp,policy",
    ("10_VPN", "Connections", "HARole"): "master,slave",
    ("07_Security", "WAFDomains", "ClientProtocol"): "HTTPS,HTTP",
    ("07_Security", "WAFDomains", "ServerProtocol"): "HTTPS,HTTP",
}

# Union sources reused by many FK columns: any account name defined in M1.
_ACCOUNTS = (("01_Foundation", "CoreAccounts", "Name"),
             ("01_Foundation", "WorkloadAccounts", "Name"))

# Foreign-key dropdowns: column -> the column whose entered values it
# references. Value is one (sheet, table, column) source, or a tuple of
# sources (union). Excel data-validation is emitted for single-source
# entries only (one DV = one range); the web UI resolves all of them.
FK_COLS = {
    ("05_Network", "HubSubnets", "VPCName"): ("05_Network", "HubVPCs", "VPCName"),
    ("05_Network", "HubERAttachments", "VPC"): ("05_Network", "HubVPCs", "VPCName"),
    ("05_Network", "SpokeSubnets", "VPCName"): ("05_Network", "SpokeVPCs", "VPCName"),
    ("05_Network", "SpokeERAttachments", "VPC"): ("05_Network", "SpokeVPCs", "VPCName"),
    ("05_Network", "SNATRules", "NATName"): ("05_Network", "NATGateways", "Name"),
    ("05_Network", "DNATRules", "NATName"): ("05_Network", "NATGateways", "Name"),
    ("05_Network", "SNATRules", "EIP"): ("05_Network", "EIPs", "Name"),
    ("05_Network", "DNATRules", "EIP"): ("05_Network", "EIPs", "Name"),
    ("05_Network", "ELBs", "VPC"): ("05_Network", "HubVPCs", "VPCName"),
    ("08_DNS", "ResolverRules", "Endpoint"): ("08_DNS", "ResolverEndpoints", "Name"),
    ("10_VPN", "Connections", "Gateway"): ("10_VPN", "Gateways", "Name"),
    ("10_VPN", "Connections", "CustomerGateway"): ("10_VPN", "CustomerGateways", "Name"),
    ("07_Security", "AntiDDoS", "EIP"): ("05_Network", "EIPs", "Name"),
    # ── mapping columns the map was missing (audit 2026-07-26) ──
    ("05_Network", "HubERAttachments", "Subnet"): ("05_Network", "HubSubnets", "Name"),
    ("05_Network", "NATGateways", "VPC"): ("05_Network", "HubVPCs", "VPCName"),
    ("05_Network", "NATGateways", "Subnet"): ("05_Network", "HubSubnets", "Name"),
    ("05_Network", "ELBs", "FrontendSubnet"): ("05_Network", "HubSubnets", "Name"),
    ("05_Network", "ELBs", "BackendSubnet"): ("05_Network", "HubSubnets", "Name"),
    ("05_Network", "ELBs", "EIP"): ("05_Network", "EIPs", "Name"),
    ("05_Network", "SpokeVPCs", "AccountName"): _ACCOUNTS,
    ("05_Network", "SpokeERAttachments", "Subnet"): ("05_Network", "SpokeSubnets", "Name"),
    ("01_Foundation", "TrustedServices", "DelegatedAdmin"): _ACCOUNTS,
    ("01_Foundation", "WorkloadAccounts", "OU"): ("01_Foundation", "OrganizationalUnits", "Name"),
    ("01_Foundation", "CoreAccounts", "OU"): ("01_Foundation", "OrganizationalUnits", "Name"),
    ("03_Identity", "AccountAssignments", "Group"): ("03_Identity", "Groups", "Name"),
    ("03_Identity", "AccountAssignments", "PermissionSet"): ("03_Identity", "PermissionSets", "Name"),
    ("03_Identity", "AccountAssignments", "AccountName"): _ACCOUNTS,
    ("03_Identity", "AppPermissionSets", "Account"): _ACCOUNTS,
    ("08_DNS", "RecordSets", "Zone"): (("08_DNS", "PublicZones", "Name"),
                                       ("08_DNS", "PrivateZones", "Name")),
    ("08_DNS", "ResolverEndpoints", "VPC"): (("05_Network", "HubVPCs", "VPCName"),
                                             ("05_Network", "SpokeVPCs", "VPCName")),
    ("11_SGACL", "SGRules", "SG"): ("11_SGACL", "SecurityGroups", "Name"),
    ("11_SGACL", "SecurityGroups", "Account"): _ACCOUNTS,
    ("10_VPN", "Gateways", "VPC"): ("05_Network", "HubVPCs", "VPCName"),
    ("10_VPN", "Gateways", "ConnectSubnet"): ("05_Network", "HubSubnets", "Name"),
    ("06_Observability", "LogConverge", "Account"): _ACCOUNTS,
}

# Scalar fields with FK semantics (web UI dropdowns; Excel scalar cells are
# not covered by _COL_POS, so no Excel DV). Key: (sheet, table, field).
SCALAR_FK = {
    ("05_Network", "Settings", "hub_account"): _ACCOUNTS,
    ("05_Network", "Settings", "snat_vpc_attachment"): ("05_Network", "HubERAttachments", "Name"),
    ("05_Network", "Settings", "enterprise_project_name"): ("02_Finance", "CostCenters", "Name"),
    ("08_DNS", "Settings", "dns_account"): _ACCOUNTS,
    ("09_CFW", "Settings", "cfw_account"): _ACCOUNTS,
    ("09_CFW", "Settings", "enterprise_project_name"): ("02_Finance", "CostCenters", "Name"),
    ("10_VPN", "Settings", "vpn_account"): _ACCOUNTS,
    ("07_Security", "Settings", "security_account"): _ACCOUNTS,
    ("06_Observability", "AuditSettings", "cts_admin_account"): _ACCOUNTS,
    ("07_Security", "WAF", "waf_vpc"): ("05_Network", "HubVPCs", "VPCName"),
    ("07_Security", "WAF", "waf_subnet"): ("05_Network", "HubSubnets", "Name"),
}

# csv-list columns whose tokens reference other tables: the web UI shows a
# live hint line of valid values (input stays free text). Each source is
# (sheet, table, column) or (sheet, table, column, prefix) — prefix is the
# token syntax to show (e.g. 'addrgroup:').
MULTI_FK = {
    ("02_Finance", "CostCenters", "Accounts"): [_ACCOUNTS[0], _ACCOUNTS[1]],
    ("03_Identity", "AppPermissionSets", "EnterpriseProjects"): [("02_Finance", "CostCenters", "Name")],
    ("06_Observability", "OpsSettings", "accounts"): [_ACCOUNTS[0], _ACCOUNTS[1]],
    ("06_Observability", "AuditSettings", "cts_no_transfer_accounts"): [_ACCOUNTS[0], _ACCOUNTS[1]],
    ("08_DNS", "PrivateZones", "VPCs"): [("05_Network", "HubVPCs", "VPCName"),
                                         ("05_Network", "SpokeVPCs", "VPCName")],
    ("08_DNS", "ResolverRules", "VPCs"): [("05_Network", "HubVPCs", "VPCName"),
                                          ("05_Network", "SpokeVPCs", "VPCName")],
    ("08_DNS", "AccessLogs", "VPCs"): [("05_Network", "HubVPCs", "VPCName"),
                                       ("05_Network", "SpokeVPCs", "VPCName")],
    ("09_CFW", "ACLRules", "Source"): [("09_CFW", "AddressGroups", "Name", "addrgroup:")],
    ("09_CFW", "ACLRules", "Destination"): [("09_CFW", "AddressGroups", "Name", "addrgroup:"),
                                            ("09_CFW", "DomainGroups", "Name", "domaingroup:")],
    ("09_CFW", "ACLRules", "Service"): [("09_CFW", "ServiceGroups", "Name", "svcgroup:")],
    ("11_SGACL", "SGRules", "Remote"): [("11_SGACL", "SecurityGroups", "Name", "sg:")],
}

# (sheet, table, column) -> (ws_title, col_letter, first_row, last_row);
# recorded while emitting, consumed by _add_validations afterwards.
_COL_POS = {}


def _set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_sheet_header(ws, sheet: Sheet):
    ws.cell(row=1, column=1, value=sheet.name).font = TITLE_FONT
    ws.cell(row=1, column=1).fill = TITLE_FILL
    ws.cell(row=1, column=1).alignment = LEFT_WRAP
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws.cell(row=2, column=1, value=sheet.description).font = NOTE_FONT
    ws.cell(row=2, column=1).fill = NOTE_FILL
    ws.cell(row=2, column=1).alignment = LEFT_WRAP
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 30


def _write_table_title(ws, row, table: Table):
    sentinel = f"### {table.name}"
    cell = ws.cell(row=row, column=1, value=sentinel)
    cell.font = SENTINEL_FONT
    cell.fill = SENTINEL_FILL
    cell.alignment = LEFT_WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    if table.description:
        ws.cell(row=row + 1, column=1, value=table.description).font = NOTE_FONT
        ws.cell(row=row + 1, column=1).fill = NOTE_FILL
        ws.cell(row=row + 1, column=1).alignment = LEFT_WRAP
        ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=8)
        return row + 2
    return row + 1


def _emit_scalar(ws, start_row, table: Table, prefill=False):
    headers = ["Field", "Type", "Default", "Sample", "Description", "Value"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=start_row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BOX
    r = start_row + 1
    for kv in table.rows:
        ws.cell(row=r, column=1, value=kv.name).font = CELL_FONT
        ws.cell(row=r, column=2, value=kv.type).font = CELL_FONT
        ws.cell(row=r, column=3, value=_repr(kv.default)).font = CELL_FONT
        sample_cell = ws.cell(row=r, column=4, value=_repr(kv.sample))
        sample_cell.font = CELL_FONT
        sample_cell.fill = SAMPLE_FILL
        ws.cell(row=r, column=5, value=kv.description).font = CELL_FONT
        # _meta ships with its Values already set (schema_version stamp).
        value_cell = ws.cell(row=r, column=6,
                             value=(_repr(kv.default) or None) if prefill else None)
        value_cell.fill = VALUE_FILL
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = BOX
            ws.cell(row=r, column=col).alignment = LEFT_WRAP
        r += 1
    return r + 1  # blank row after


def _emit_list_single(ws, start_row, table: Table):
    headers = ["Value"]
    c = ws.cell(row=start_row, column=1, value="Value")
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = BOX
    sample_header = ws.cell(row=start_row, column=2, value="Sample")
    sample_header.font = HEADER_FONT
    sample_header.fill = HEADER_FILL
    sample_header.alignment = CENTER
    sample_header.border = BOX

    r = start_row + 1
    # Pre-populate Value column with sample values (user can clear or edit)
    n_rows = max(len(table.sample_rows), 8)
    for i in range(n_rows):
        val = table.sample_rows[i] if i < len(table.sample_rows) else None
        # Place sample in column 2 (read-only-ish reference)
        if i < len(table.sample_rows):
            s = ws.cell(row=r, column=2, value=val)
            s.fill = SAMPLE_FILL
            s.font = CELL_FONT
            s.border = BOX
        else:
            ws.cell(row=r, column=2).border = BOX
        v = ws.cell(row=r, column=1, value=val if i < len(table.sample_rows) else None)
        v.fill = VALUE_FILL
        v.font = CELL_FONT
        v.border = BOX
        r += 1
    return r + 1


def _emit_object_table(ws, start_row, table: Table):
    has_toggle = not table.mandatory and not _table_has_explicit_enabled(table)
    # Build column list
    cols = []
    if has_toggle:
        cols.append(("Enabled", "bool", "TRUE = include this row in the deployment"))
    cols.extend(table.columns)

    # Header
    for col_idx, (name, typ, desc) in enumerate(cols, 1):
        c = ws.cell(row=start_row, column=col_idx, value=name)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BOX
    # Type row (small grey row showing types)
    for col_idx, (name, typ, desc) in enumerate(cols, 1):
        c = ws.cell(row=start_row + 1, column=col_idx, value=f"({typ})")
        c.font = NOTE_FONT
        c.fill = NOTE_FILL
        c.alignment = CENTER
        c.border = BOX

    r = start_row + 2
    # Sample / pre-filled rows (in Value-coloured cells so user knows they're editable)
    n_blanks = 8
    for sample in table.sample_rows:
        for col_idx, (name, typ, desc) in enumerate(cols, 1):
            val = sample.get(name)
            if name == "Enabled" and val is None:
                val = True
            cell = ws.cell(row=r, column=col_idx, value=_cell_value(val))
            cell.fill = VALUE_FILL
            cell.font = CELL_FONT
            cell.border = BOX
            cell.alignment = LEFT_WRAP
        r += 1
    # Blank fillable rows
    for _ in range(n_blanks):
        for col_idx, (name, typ, desc) in enumerate(cols, 1):
            cell = ws.cell(row=r, column=col_idx, value=None)
            cell.fill = VALUE_FILL
            cell.font = CELL_FONT
            cell.border = BOX
        r += 1
    last_row = r - 1
    # Record positions
    for col_idx, (name, typ, desc) in enumerate(cols, 1):
        _COL_POS[(ws.title, table.name, name, "type")] = typ
        _COL_POS[(ws.title, table.name, name)] = (
            ws.title, get_column_letter(col_idx), start_row + 2, last_row)
    return r + 1


def _table_has_explicit_enabled(table: Table) -> bool:
    return any(c[0] == "Enabled" for c in table.columns)


def _repr(v):
    if v is None:
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, list):
        return ",".join(str(x) for x in v)
    return str(v)


def _cell_value(v):
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, list):
        return ",".join(str(x) for x in v)
    return v


def _emit_sheet(wb, sheet: Sheet):
    ws = wb.create_sheet(sheet.name)
    _write_sheet_header(ws, sheet)

    # Column widths — uniform default; first column wider for field names
    _set_widths(ws, [28, 14, 22, 28, 50, 28, 18, 18])

    row = 4
    for table in sheet.tables:
        row = _write_table_title(ws, row, table)
        if table.kind == "scalar":
            row = _emit_scalar(ws, row, table, prefill=(sheet.name == "_meta"))
        elif table.kind == "list-single":
            row = _emit_list_single(ws, row, table)
        elif table.kind == "object-table":
            row = _emit_object_table(ws, row, table)
        else:
            raise ValueError(f"unknown table kind: {table.kind}")

    ws.freeze_panes = "A4"


def _add_validations(wb):
    """Dropdowns: TRUE/FALSE on every bool column, static enums, and
    FK pick-lists referencing the range where the target names are typed."""
    n_dv = 0
    by_sheet = {}

    def _dv_for(key, formula):
        nonlocal n_dv
        pos = _COL_POS.get(key)
        if not pos:
            return
        ws_title, col, first, last = pos
        ws = wb[ws_title]
        dv = DataValidation(type="list", formula1=formula, allow_blank=True,
                            showErrorMessage=True)
        ws.add_data_validation(dv)
        dv.add(f"{col}{first}:{col}{last}")
        n_dv += 1

    seen = set()
    for key, val in list(_COL_POS.items()):
        if len(key) == 4 or key in seen:
            continue
        sheet, table, colname = key
        seen.add(key)
        typ = _COL_POS.get((sheet, table, colname, "type"))
        if typ == "bool":
            _dv_for(key, '"TRUE,FALSE"')
        elif (sheet, table, colname) in ENUM_COLS:
            _dv_for(key, f'"{ENUM_COLS[(sheet, table, colname)]}"')
        elif (sheet, table, colname) in FK_COLS:
            val = FK_COLS[(sheet, table, colname)]
            srcs = val if isinstance(val[0], tuple) else (val,)
            if len(srcs) == 1:  # Excel DV takes one range; unions are UI-only
                pos = _COL_POS.get(srcs[0])
                if pos:
                    s_ws, s_col, s_first, s_last = pos
                    _dv_for(key, f"='{s_ws}'!${s_col}${s_first}:${s_col}${s_last}")
    return n_dv


def main():
    if len(sys.argv) != 2:
        print("usage: python gen_template.py <output.xlsx>")
        sys.exit(1)
    out = Path(sys.argv[1])
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet

    _COL_POS.clear()
    for sheet in SHEETS:
        _emit_sheet(wb, sheet)
    n_dv = _add_validations(wb)

    wb.save(out)
    print(f"Wrote {out} ({n_dv} dropdown validations)")


if __name__ == "__main__":
    main()
