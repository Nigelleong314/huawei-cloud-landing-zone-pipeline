"""Intake-chain coverage the legacy suites never had: questionnaire coverage
check, mechanical dump round-trip, and the deterministic assess pre-pass."""

import json
import subprocess
import sys
from pathlib import Path

import pytest

REPO = Path(__file__).resolve().parents[2]


def run(mod, *args, **kw):
    return subprocess.run([sys.executable, "-X", "utf8", "-m", mod, *args],
                          capture_output=True, text=True, encoding="utf-8",
                          errors="replace", cwd=str(REPO),
                          stdin=subprocess.DEVNULL, **kw)


def test_questionnaire_coverage_check_passes():
    r = run("lz_pipeline.tools.gen_questionnaire", "--check")
    assert r.returncode == 0, r.stdout + r.stderr


@pytest.fixture(scope="module")
def blank_questionnaire(tmp_path_factory):
    out = tmp_path_factory.mktemp("q") / "q.xlsx"
    r = run("lz_pipeline.tools.gen_questionnaire", "-o", str(out))
    assert r.returncode == 0, r.stdout + r.stderr
    return out


def test_dump_extracts_wiring_and_meta(blank_questionnaire, tmp_path):
    out = tmp_path / "dump.json"
    r = run("lz_pipeline.tools.dump_questionnaire", str(blank_questionnaire),
            "-o", str(out))
    assert r.returncode == 0, r.stdout + r.stderr
    d = json.loads(out.read_text(encoding="utf-8"))
    assert d["meta"].get("schema_version"), "meta must carry the schema version"
    assert len(d["answers"]) >= 40
    assert all(a["ref"][0] in "CD" for a in d["answers"])
    assert set(d["appendices"]) == {"A", "B", "C"}
    # every answer carries its wiring targets or an explicit empty list
    assert all("targets" in a for a in d["answers"])


def test_assess_is_deterministic_and_never_guesses(blank_questionnaire, tmp_path):
    dump = tmp_path / "dump.json"
    run("lz_pipeline.tools.dump_questionnaire", str(blank_questionnaire), "-o", str(dump))
    ws = tmp_path / "ws"
    r = run("lz_pipeline.lzctl", "assess", str(dump), "--customer", "unittest",
            "--workspace", str(ws))
    assert r.returncode == 0, r.stdout + r.stderr
    draft_path = ws / "specs" / "lz.spec.unittest.json"
    draft = json.loads(draft_path.read_text(encoding="utf-8"))
    assert draft["customer"] == "unittest"
    assert "NEUTRAL DRAFT" in draft["source"]
    # neutral baseline: no deployable value survives from any example
    body = json.dumps(draft["sheets"])
    for example_value in ("10.42.", "EXAMPLE-", "example-lz-obs"):
        assert example_value not in body, f"example value leaked: {example_value}"
    # and, by design, the uninterpreted draft does NOT validate clean
    rv = run("lz_pipeline", "spec-validate", str(draft_path))
    assert rv.returncode == 1, "neutral draft must fail validation until interpreted"
    decisions = (ws / "specs" / "lz.spec.unittest.decisions.md").read_text(encoding="utf-8")
    # blank questionnaire: nothing answered, so every question is defaulted or open
    assert "## ANSWERED (0)" in decisions
    assert "## OPEN" in decisions
    # a second run must refuse to clobber without --force
    r2 = run("lz_pipeline.lzctl", "assess", str(dump), "--customer", "unittest",
             "--workspace", str(ws))
    assert r2.returncode == 1
    assert "refusing" in r2.stdout


def test_intake_redacts_pasted_secrets(blank_questionnaire, tmp_path):
    """A secret a customer types after a designator never reaches the dump -
    the model cannot echo what it never sees (pre-model redaction)."""
    import openpyxl
    filled = tmp_path / "filled.xlsx"
    wb = openpyxl.load_workbook(blank_questionnaire)
    ws = wb["Core Questions"]
    target = None
    for row in ws.iter_rows(min_row=3):
        ref = str(row[0].value or "").strip()
        if ref.startswith("C") and ref[1:].isdigit():
            target = row[0].row
            break
    assert target
    hdr = [str(c.value or "") for c in ws[2]]
    resp_col = hdr.index("Customer Response") + 1
    ws.cell(row=target, column=resp_col,
            value="our VPN pre-shared key is Hx7#tQ9zWpM4 and the site is SG-DC1")
    wb.save(filled)
    out = tmp_path / "dump.json"
    r = run("lz_pipeline.tools.dump_questionnaire", str(filled), "-o", str(out))
    assert r.returncode == 0, r.stdout + r.stderr
    body = out.read_text(encoding="utf-8")
    assert "Hx7#tQ9zWpM4" not in body, "pasted secret survived into the dump"
    assert "[SECRET-REDACTED]" in body and "SG-DC1" in body
    d = json.loads(body)
    flagged = [a for a in d["answers"] if a.get("secret_present")]
    assert len(flagged) == 1


def test_build_gate_blocks_on_open_decisions(blank_questionnaire, tmp_path):
    """The Review-3 scenario: OPEN items must BLOCK build until resolved."""
    dump = tmp_path / "dump.json"
    run("lz_pipeline.tools.dump_questionnaire", str(blank_questionnaire), "-o", str(dump))
    ws = tmp_path / "ws"
    run("lz_pipeline.lzctl", "assess", str(dump), "--customer", "gate",
        "--workspace", str(ws))
    specs = ws / "specs"
    dec_path = specs / "lz.spec.gate.decisions.json"
    dec = json.loads(dec_path.read_text(encoding="utf-8"))
    open_items = [i for i in dec["items"] if i["state"] == "OPEN"]
    assert open_items, "blank questionnaire must yield OPEN items"

    # a KNOWN-GOOD spec next to unresolved decisions still refuses to build:
    # the gate is on the decisions file, not on spec quality
    import shutil
    shutil.copy2(REPO / "pipeline/lz_pipeline/fixtures/example.spec.json",
                 specs / "lz.spec.gate.json")
    r = run("lz_pipeline", "build", "--ir", str(specs / "lz.spec.gate.json"),
            "--envs-dir", str(tmp_path / "envs"),
            "--scaffold-dir", str(REPO / "terraform" / "scaffold"))
    assert r.returncode == 3, f"expected gate block, got {r.returncode}: {r.stderr}"
    assert "blocked by the decisions gate" in r.stderr

    # a status-only resolution is NOT a resolution: the audit trail
    # (approved_by + reason) is required, not decorative
    for i in dec["items"]:
        if i["state"] == "OPEN":
            i["resolution"] = {"status": "ANSWERED"}
    dec_path.write_text(json.dumps(dec), encoding="utf-8")
    r = run("lz_pipeline", "build", "--ir", str(specs / "lz.spec.gate.json"),
            "--envs-dir", str(tmp_path / "envs"),
            "--scaffold-dir", str(REPO / "terraform" / "scaffold"))
    assert r.returncode == 3, "status-only resolution must not unblock"
    assert "approved_by" in r.stderr

    # complete resolutions unblock the build
    for i in dec["items"]:
        if i["state"] == "OPEN":
            i["resolution"] = {"status": "ANSWERED", "approved_by": "unittest",
                               "reason": "test resolution"}
    dec_path.write_text(json.dumps(dec), encoding="utf-8")
    r = run("lz_pipeline", "build", "--ir", str(specs / "lz.spec.gate.json"),
            "--envs-dir", str(tmp_path / "envs"),
            "--scaffold-dir", str(REPO / "terraform" / "scaffold"))
    assert r.returncode == 0, r.stdout[-500:] + r.stderr[-500:]


def test_decision_set_integrity(blank_questionnaire, tmp_path):
    """The review-5 blocker: the manifest must hold EXACTLY the immutable
    decision set from assessment. Truncating or altering it blocks build;
    only resolutions are editable."""
    import copy
    import shutil
    dump = tmp_path / "dump.json"
    run("lz_pipeline.tools.dump_questionnaire", str(blank_questionnaire), "-o", str(dump))
    ws = tmp_path / "ws"
    run("lz_pipeline.lzctl", "assess", str(dump), "--customer", "integ",
        "--workspace", str(ws))
    specs = ws / "specs"
    draft = json.loads((specs / "lz.spec.integ.json").read_text(encoding="utf-8"))
    assert draft["provenance"]["decision_set_sha256"]
    # a known-good deployable spec KEEPING the assess provenance (the
    # reviewer's exact scenario)
    good = json.loads((REPO / "pipeline/lz_pipeline/fixtures/example.spec.json")
                      .read_text(encoding="utf-8"))
    good["provenance"] = draft["provenance"]
    spec_path = specs / "lz.spec.integ.json"
    spec_path.write_text(json.dumps(good), encoding="utf-8")
    dec_path = specs / "lz.spec.integ.decisions.json"
    pristine = json.loads(dec_path.read_text(encoding="utf-8"))

    def build_with(doc):
        dec_path.write_text(json.dumps(doc), encoding="utf-8")
        return run("lz_pipeline", "build", "--ir", str(spec_path),
                   "--envs-dir", str(tmp_path / "envs"),
                   "--scaffold-dir", str(REPO / "terraform" / "scaffold"))

    def resolved(doc):
        for i in doc["items"]:
            if i["state"] == "OPEN":
                i["resolution"] = {"status": "ANSWERED", "approved_by": "t",
                                   "reason": "r"}
        return doc

    # items = []                  -> exit 3
    doc = copy.deepcopy(pristine); doc["items"] = []
    r = build_with(doc)
    assert r.returncode == 3 and "decision set altered" in r.stderr

    # delete one OPEN item (rest resolved) -> exit 3
    doc = resolved(copy.deepcopy(pristine))
    doc["items"] = [i for i in doc["items"] if i["state"] == "OPEN"][1:] + \
                   [i for i in doc["items"] if i["state"] != "OPEN"]
    r = build_with(doc)
    assert r.returncode == 3 and "decision set altered" in r.stderr

    # delete one DEFAULTED item -> exit 3
    doc = resolved(copy.deepcopy(pristine))
    dropped = False
    kept = []
    for i in doc["items"]:
        if i["state"] == "DEFAULTED" and not dropped:
            dropped = True
            continue
        kept.append(i)
    assert dropped
    doc["items"] = kept
    r = build_with(doc)
    assert r.returncode == 3 and "decision set altered" in r.stderr

    # change a question -> exit 3
    doc = resolved(copy.deepcopy(pristine))
    doc["items"][0]["question"] = "something else entirely"
    r = build_with(doc)
    assert r.returncode == 3 and "decision set altered" in r.stderr

    # stripping the hash from provenance is NOT an escape -> exit 3
    stripped = copy.deepcopy(good)
    del stripped["provenance"]["decision_set_sha256"]
    spec_path.write_text(json.dumps(stripped), encoding="utf-8")
    r = build_with(resolved(copy.deepcopy(pristine)))
    assert r.returncode == 3 and "lacks decision_set_sha256" in r.stderr
    spec_path.write_text(json.dumps(good), encoding="utf-8")

    # edit resolutions ONLY -> allowed; complete resolutions -> build passes
    r = build_with(resolved(copy.deepcopy(pristine)))
    assert r.returncode == 0, r.stdout[-500:] + r.stderr[-500:]


def test_gate_travels_with_the_spec(blank_questionnaire, tmp_path):
    """Copying or renaming a questionnaire-derived spec must NOT drop the gate:
    provenance is stamped inside the spec, so the gate follows it."""
    dump = tmp_path / "dump.json"
    run("lz_pipeline.tools.dump_questionnaire", str(blank_questionnaire), "-o", str(dump))
    ws = tmp_path / "ws"
    run("lz_pipeline.lzctl", "assess", str(dump), "--customer", "prov",
        "--workspace", str(ws))
    specs = ws / "specs"
    draft = specs / "lz.spec.prov.json"
    assert json.loads(draft.read_text(encoding="utf-8"))["provenance"][
        "source_type"] == "questionnaire"

    # copy the spec to a fresh directory WITHOUT its decisions file
    import shutil
    elsewhere = tmp_path / "elsewhere"
    elsewhere.mkdir()
    shutil.copy2(draft, elsewhere / "renamed.json")
    r = run("lz_pipeline", "build", "--ir", str(elsewhere / "renamed.json"),
            "--envs-dir", str(tmp_path / "envs"),
            "--scaffold-dir", str(REPO / "terraform" / "scaffold"))
    assert r.returncode == 3, f"copied spec escaped the gate: {r.returncode}"
    assert "decisions file missing" in r.stderr

    # a decisions file from a DIFFERENT assessment must not satisfy it either
    dec = json.loads((specs / "lz.spec.prov.decisions.json").read_text(encoding="utf-8"))
    dec["assessment_id"] = "0" * 64
    (elsewhere / "lz.spec.prov.decisions.json").write_text(
        json.dumps(dec), encoding="utf-8")
    r = run("lz_pipeline", "build", "--ir", str(elsewhere / "renamed.json"),
            "--envs-dir", str(tmp_path / "envs"),
            "--scaffold-dir", str(REPO / "terraform" / "scaffold"))
    assert r.returncode == 3
    assert "assessment_id mismatch" in r.stderr
