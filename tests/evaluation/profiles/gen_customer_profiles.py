"""Generate filled assessment questionnaires for validation runs.

Ten synthetic customers, deliberately UNEVEN: a corpus where every workbook
is equally complete tests one path ten times. Completeness is a dial per
profile (`depth`), and the facts a customer withholds are chosen to land on
known pipeline seams:

  thorough  - appendices full, CIDRs and email pattern given, VPN detail
              supplied. Should reach a near-clean spec.
  moderate  - appendices present but partial; some implementation values
              described in prose instead of stated.
  sparse    - narrative answers, thin appendices, several deep-dive
              questions left blank. Exercises DEFAULTED/OPEN classification
              and the gap-registration path.

No real organisation, person, address, or IP range appears here: names are
invented, domains use .example, and every CIDR is RFC1918.

Run:  py tests/evaluation/profiles/gen_customer_profiles.py -o <dir>
"""

import argparse
import subprocess
import sys
from pathlib import Path

REPO = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(REPO / "pipeline"))

from openpyxl import load_workbook

# ── profiles ────────────────────────────────────────────────────────────────
# supernet: the private range the customer can hand to Huawei Cloud.
# waves:    (application, environment) pairs for the first migration wave.
# teams:    (team, responsibility, access) triples.

PROFILES = [
    dict(slug="summit_retail", name="Summit Retail Group", country="au",
         source="aws", region="ap-southeast-3", depth="thorough",
         industry="retail", domain="summitretailgrp.example",
         supernet="10.60.0.0/16", accounts=22, dc="Sydney and Melbourne",
         regs="PCI DSS for card handling, Australian Privacy Principles",
         waves=[("Ecommerce Storefront", "prod"), ("Order Management", "prod"),
                ("Loyalty Platform", "prod"), ("Merchandising Analytics", "uat")],
         teams=[("Cloud Platform", "Cloud foundation and networking", "admin"),
                ("Digital Commerce", "Storefront and order services", "custom"),
                ("Security Operations", "Security monitoring and response", "custom"),
                ("Data & Analytics", "Reporting and merchandising models", "readonly")],
         siem="Splunk Cloud", idp="Microsoft Entra ID", cicd="GitHub Actions"),

    dict(slug="keris_telecom", name="Keris Telecom Berhad", country="my",
         source="onprem", region="ap-southeast-3", depth="moderate",
         industry="telecommunications", domain="keristelecom.example",
         supernet="10.72.0.0/16", accounts=14, dc="Kuala Lumpur and Cyberjaya",
         regs="MCMC licence conditions, Malaysian PDPA",
         waves=[("Subscriber Portal", "prod"), ("Billing Mediation", "prod"),
                ("Network Analytics", "nonprod")],
         teams=[("Cloud Infrastructure", "Cloud platform and connectivity", "admin"),
                ("BSS Engineering", "Billing and subscriber systems", "custom"),
                ("Information Security", "Security and compliance", "custom")],
         siem="IBM QRadar", idp="Active Directory with AD FS", cicd="GitLab CI"),

    dict(slug="sakura_media", name="Sakura Media Holdings", country="jp",
         source="gcp", region="ap-southeast-1", depth="thorough",
         industry="media and broadcasting", domain="sakuramediahd.example",
         supernet="10.84.0.0/16", accounts=19, dc="Tokyo (colocation)",
         regs="APPI, broadcast content retention obligations",
         waves=[("Streaming Delivery", "prod"), ("Content Management", "prod"),
                ("Ad Decisioning", "prod"), ("Media Transcode Farm", "uat")],
         teams=[("Platform Engineering", "Cloud foundation", "admin"),
                ("Streaming Services", "Delivery and playback", "custom"),
                ("Security", "Security operations", "custom"),
                ("Content Operations", "Media pipeline operations", "readonly")],
         siem="Microsoft Sentinel", idp="Okta", cicd="GitHub Actions"),

    dict(slug="delta_agri", name="Delta Agritech Vietnam", country="vn",
         source="onprem", region="ap-southeast-3", depth="sparse",
         industry="agriculture technology", domain="deltaagrivn.example",
         supernet="", accounts=8, dc="Can Tho",
         regs="Vietnam Cybersecurity Law data localisation",
         waves=[("Farm Telemetry", "prod"), ("Supply Chain Portal", "nonprod")],
         teams=[("IT Operations", "All infrastructure", "admin"),
                ("Application Team", "Farm and supply applications", "custom")],
         siem="", idp="", cicd="Jenkins"),

    dict(slug="orion_edu", name="Orion Education Network", country="ph",
         source="azure", region="ap-southeast-3", depth="moderate",
         industry="higher education", domain="orionedunet.example",
         supernet="10.96.0.0/16", accounts=12, dc="Manila campus",
         regs="Philippine Data Privacy Act, student record retention",
         waves=[("Student Information System", "prod"), ("Learning Platform", "prod"),
                ("Research Computing", "nonprod")],
         teams=[("Central IT", "Cloud platform and identity", "admin"),
                ("Academic Systems", "Student and learning systems", "custom"),
                ("Research Support", "Research computing", "readonly")],
         siem="", idp="Microsoft Entra ID", cicd="Azure DevOps"),

    dict(slug="bluereef_hotels", name="BlueReef Hotels & Resorts", country="id",
         source="aws", region="ap-southeast-3", depth="moderate",
         industry="hospitality", domain="bluereefhotels.example",
         supernet="10.108.0.0/16", accounts=16, dc="Jakarta and Bali",
         regs="PCI DSS for payments, Indonesian PDP Law",
         waves=[("Booking Engine", "prod"), ("Property Management", "prod"),
                ("Guest Mobile API", "prod"), ("Revenue Analytics", "uat")],
         teams=[("Cloud Team", "Cloud platform", "admin"),
                ("Digital Booking", "Booking and guest services", "custom"),
                ("Security & Compliance", "Security and PCI scope", "custom")],
         siem="Splunk Enterprise", idp="Microsoft Entra ID", cicd="Bitbucket Pipelines"),

    dict(slug="ironclad_mfg", name="Ironclad Manufacturing", country="th",
         source="vmware", region="ap-southeast-3", depth="sparse",
         industry="industrial manufacturing", domain="ironcladmfg.example",
         supernet="10.120.0.0/16", accounts=9, dc="Rayong plant and Bangkok office",
         regs="Thailand PDPA; OT network separation required",
         waves=[("MES Reporting", "prod"), ("Plant Data Historian", "nonprod")],
         teams=[("IT Infrastructure", "Cloud and on-premises infrastructure", "admin"),
                ("Manufacturing Systems", "Plant applications", "custom")],
         siem="", idp="Active Directory with AD FS", cicd=""),

    dict(slug="lotus_pharma", name="Lotus Pharmaceuticals", country="in",
         source="azure", region="ap-southeast-1", depth="thorough",
         industry="pharmaceutical manufacturing", domain="lotuspharma.example",
         supernet="10.132.0.0/16", accounts=24, dc="Hyderabad and Pune",
         regs="GxP validation, 21 CFR Part 11, India DPDP Act",
         waves=[("Clinical Data Platform", "prod"), ("Manufacturing Quality", "prod"),
                ("Regulatory Submissions", "prod"), ("Research Analytics", "uat")],
         teams=[("Cloud Platform", "Cloud foundation and validation", "admin"),
                ("Clinical Systems", "Clinical data applications", "custom"),
                ("Quality Assurance", "GxP validation and audit", "readonly"),
                ("Security", "Security operations", "custom")],
         siem="Microsoft Sentinel", idp="Microsoft Entra ID", cicd="Azure DevOps"),

    dict(slug="zenith_capital", name="Zenith Capital Partners", country="hk",
         source="multicloud", region="ap-southeast-1", depth="thorough",
         industry="asset management", domain="zenithcapitalhk.example",
         supernet="10.144.0.0/16", accounts=20, dc="Hong Kong and Singapore",
         regs="HKMA SA-2 outsourcing, SFC electronic trading, seven-year retention",
         waves=[("Portfolio Analytics", "prod"), ("Client Reporting", "prod"),
                ("Trade Surveillance", "prod"), ("Quant Research", "uat")],
         teams=[("Cloud Engineering", "Cloud platform", "admin"),
                ("Investment Technology", "Portfolio and trading systems", "custom"),
                ("Risk & Compliance", "Surveillance and reporting", "readonly"),
                ("Information Security", "Security operations", "custom")],
         siem="Splunk Cloud", idp="Okta", cicd="GitLab CI"),

    dict(slug="terrafirma_mining", name="TerraFirma Mining", country="mn",
         source="oci", region="ap-southeast-3", depth="sparse",
         industry="mining and resources", domain="terrafirmamining.example",
         supernet="", accounts=7, dc="Ulaanbaatar office; remote site links",
         regs="Mongolian data residency guidance; site safety reporting",
         waves=[("Fleet Telemetry", "prod"), ("Geology Data Store", "nonprod")],
         teams=[("IT Services", "All cloud and on-site IT", "admin"),
                ("Mine Systems", "Fleet and geology applications", "custom")],
         siem="", idp="", cicd=""),
]


# ── answer composition ──────────────────────────────────────────────────────

def _acct(p, app, env):
    stem = "-".join(app.lower().split()[:2])[:20]
    return f"hc-{stem}-{env}"


def answers(p):
    """{ref: response}. A blank value means the customer left it empty."""
    d, deep = p["depth"], p["depth"] != "sparse"
    thorough = d == "thorough"
    apps = ", ".join(a for a, _ in p["waves"])
    a = {}

    a["C1"] = (f"We run roughly {p['accounts']} accounts on "
               f"{ {'aws':'AWS','azure':'Microsoft Azure','gcp':'Google Cloud','oci':'Oracle Cloud','vmware':'VMware on-premises','onprem':'our own data centres','multicloud':'AWS and Azure'}[p['source']] }, "
               f"plus {p['dc']}. This Huawei Cloud landing zone is for our "
               f"{p['industry']} workloads in {p['country'].upper()}.")
    a["C2"] = ("No production workloads on Huawei Cloud today. We want a clean "
               "landing zone and will only adopt an existing account if a "
               "migration dependency forces it.")
    a["C3"] = (f"First wave: {apps}. Production and non-production must be "
               f"separate accounts; the exact split can follow your reference "
               f"design. See Appendix A." if thorough else
               f"First wave: {apps}. We want production isolated from everything else.")
    a["C4"] = ("Yes - we use a management-group hierarchy today and would like an "
               "equivalent OU structure." if p["source"] in ("azure", "multicloud")
               else "Not formally. We group by project and would like a proper OU design.")
    a["C5"] = (f"Primary region {p['region']}. No secondary region in this phase; "
               f"we will revisit DR after the first wave is live." if not thorough
               else f"Primary region {p['region']}. We expect a secondary region for "
                    f"disaster recovery within 12 months but it is out of scope now.")
    a["C6"] = (f"Use huawei-<account>@{p['domain']} - the mailbox alias is already "
               f"reserved and forwards to the cloud team."
               if d != "sparse" else
               "We have not agreed a mailbox pattern yet; our messaging team must confirm.")
    a["C7"] = (f"Yes: <org>-<country>-<env>-<service>-<nn>, lowercase, hyphenated. "
               f"Org prefix is '{p['slug'].split('_')[0]}'." if d != "sparse"
               else "Nothing formal - please propose one and we will adopt it.")

    a["C8"] = "See Appendix C for teams and responsibilities."
    a["C9"] = (f"Federated sign-in through {p['idp']}." if p["idp"]
               else "Undecided. Today we use local accounts and would like advice.")
    a["C10"] = ("Yes, SAML 2.0 and SCIM are both supported and already in use for "
                "other clouds." if p["idp"] else "")
    a["C11"] = ("Yes - a managed service partner needs scoped operational access, "
                "and external auditors need read-only access at review time."
                if d != "sparse" else "Possibly later; nothing confirmed.")
    a["C12"] = ("MFA for all human access, 14-character minimum passwords, 90-day "
                "rotation for privileged accounts, lockout after 5 failures."
                if d != "sparse" else "Follow your recommended baseline.")

    a["C13"] = (f"{p['dc']} connected over MPLS, with internet breakout at the main "
                f"site. Full route inventory available on request.")
    a["C14"] = ("Site-to-site VPN first, with Direct Connect once volumes justify it."
                if d != "sparse" else
                "We need connectivity to our sites but have not chosen a method.")
    a["C15"] = ("Outbound internet via a shared NAT path with egress inspection; only "
                "the published applications accept inbound traffic.")
    a["C16"] = ("Normal enterprise expectations - no special latency or bandwidth "
                "constraints between VPCs." if d != "sparse" else "")
    a["C17"] = (f"Yes - {p['supernet']} is reserved for Huawei Cloud and does not "
                f"overlap anything in use. See Appendix B." if p["supernet"] else
                "Not yet. Our network team must confirm a free range before we can "
                "commit; existing addressing is fragmented.")
    a["C18"] = ("One VPC per workload account, with shared connectivity in a hub."
                if d != "sparse" else "Please propose a structure.")
    a["C19"] = (f"{p['waves'][0][0]} must reach shared services and the data tier; "
                f"non-production must not reach production.")
    a["C20"] = ("Mandatory tags: application, owner, environment, costcentre."
                if d != "sparse" else "We tag inconsistently today; propose a set.")

    a["C21"] = (f"{p['siem']} is our primary security tooling and must receive "
                f"Huawei Cloud logs." if p["siem"] else
                "Limited tooling today - endpoint protection and firewall logs only.")
    a["C22"] = ("Cloud Firewall with intrusion prevention enabled at the perimeter, "
                "east-west inspection between production and non-production."
                if d != "sparse" else "We expect a managed firewall; details to follow.")
    a["C23"] = ("SecMaster and Host Security are of interest but not committed for "
                "this phase - treat them as future scope." if d != "sparse" else "")
    a["C24"] = ("Yes: no disabling of audit logging, no deleting log buckets, no "
                "creating IAM users outside the approved federation, and deployment "
                f"restricted to {p['region']}.")

    a["C25"] = (f"All account activity centralised into a protected, immutable audit "
                f"store owned by security. Retention {'seven years' if 'seven' in p['regs'] else 'at least one year'}.")
    a["C26"] = p["regs"]
    a["C27"] = ("Yes - we use continuous compliance controls in our current cloud and "
                "expect an equivalent here, owned by the security team."
                if d != "sparse" else "Not today.")

    a["C28"] = (f"{p['siem']} plus platform-native monitoring." if p["siem"]
                else "Basic infrastructure monitoring; we want to improve this.")
    a["C29"] = ("Yes - centralise audit, network flow, and application logs into a "
                "dedicated logging account.")
    a["C30"] = ("Searchable 90 days, archived to cold storage after that; audit logs "
                "kept per the retention obligation above." if d != "sparse" else
                "Follow your default retention and we will review it.")

    a["C31"] = ("Finance owns the cloud budget; the cloud platform team owns technical "
                "cost control. Both need spend visibility per account.")
    a["C32"] = ("Charged back to business units by tag and account." if d != "sparse"
                else "Not decided.")
    a["C33"] = ("Group by application within each account so cost reports match our "
                "internal structure." if thorough else "")

    # ── deep-dive ────────────────────────────────────────────────────────
    a["D1"] = (f"Growth to roughly {p['accounts'] + 10} accounts over two years as "
               f"further waves migrate. The cloud platform team approves new accounts.")
    a["D2"] = (f"{p['cicd']}, running in our own environment. Pipelines should use "
               f"short-lived federated credentials, not static keys." if p["cicd"] else "")
    a["D3"] = ("Two break-glass accounts held in a sealed process, MFA-protected, "
               "reviewed quarterly." if deep else "")
    a["D4"] = ("Admin, power-user, read-only, plus a security-audit role." if deep else "")
    a["D5"] = ("Two VPN tunnels from our primary data centre firewalls, BGP routing. "
               "Device models, public IPs, and ASN will be supplied by the network "
               "team before implementation." if deep else "")
    a["D6"] = ""
    a["D7"] = ("Separate subnets per tier - web, application, data - in every VPC."
               if thorough else "")
    a["D8"] = ("Yes, container workloads are planned for the application tier in a "
               "later wave." if deep else "")
    a["D9"] = (f"Conditional forwarding between Huawei Cloud and our internal DNS. "
               f"Resolver IPs to be confirmed by the network team." if deep else "")
    a["D10"] = (f"{p['waves'][0][0]} is internet-facing with TLS terminating at the "
                f"load balancer." if deep else "")
    a["D11"] = ""
    a["D12"] = ("Yes - flow logs retained 90 days for investigation." if thorough else "")
    a["D13"] = (f"{p['waves'][0][0]} needs Web Application Firewall protection."
                if deep else "")
    a["D14"] = ""
    a["D15"] = ("We maintain a geo-blocking policy and a threat feed we would like "
                "applied at the perimeter." if thorough else "")
    a["D16"] = ("Keys managed in Huawei Cloud KMS with rotation; no external HSM "
                "requirement." if deep else "")
    a["D17"] = ("No. No object storage should be publicly readable." if deep else "")
    a["D18"] = ""
    a["D19"] = ("Platform alerts to the cloud team distribution list, security alerts "
                "to the security team. Email is sufficient initially; the exact "
                "addresses will be confirmed before go-live." if deep else "")
    a["D20"] = (f"Yes - {p['siem']}. Audit, firewall, and DNS logs should be forwarded."
                if p["siem"] else "")
    a["D21"] = ("Daily backups with 30-day retention for production; recovery point "
                "objective of 24 hours is acceptable in this phase." if deep else "")
    a["D22"] = ("Shared platform costs split across business units by consumption."
                if thorough else "")
    a["D23"] = ""
    return a


def appendix_a(p):
    rows = []
    for app, env in p["waves"]:
        stem = "-".join(app.lower().split()[:2])[:13]
        rows.append([app, env, _acct(p, app, env),
                     f"Workloads/{'Prod' if env == 'prod' else 'NonProd'}",
                     f"{stem}@{p['domain']}",
                     "internet-facing" if app == p["waves"][0][0] else "private"])
    return rows if p["depth"] != "sparse" else rows[:1]


def appendix_b(p):
    if not p["supernet"]:
        return [["Available supernet", "", "Huawei Cloud allocation", "All",
                 "Network team must confirm a non-overlapping range"]]
    base = p["supernet"].split(".")[0] + "." + p["supernet"].split(".")[1]
    rows = [["Available supernet", p["supernet"], "Huawei Cloud allocation", "All",
             "Reserved; subject to final overlap validation"],
            ["Planned", f"{base}.0.0/20", "Network hub / shared connectivity",
             "Platform", "Initial reservation"],
            ["Avoid", "10.0.0.0/8", "Existing enterprise networks", "-",
             "Detailed route inventory available on request"]]
    if p["depth"] == "thorough":
        for i, (app, env) in enumerate(p["waves"]):
            rows.append(["Planned", f"{base}.{16 + i * 16}.0/20", f"{app} VPC",
                         _acct(p, app, env), "Per-workload allocation"])
    return rows


def appendix_c(p):
    return [[t, resp, f"{t.lower().replace(' ', '-').replace('&', 'and')}@{p['domain']}",
             acc, "Relevant platform or workload accounts",
             "Federated access preferred" if p["idp"] else "Access method to be agreed"]
            for t, resp, acc in p["teams"]]


# ── workbook emit ───────────────────────────────────────────────────────────

def fill(blank: Path, out: Path, p: dict):
    wb = load_workbook(blank)
    ans = answers(p)

    for sheet in ("Core Questions", "Deep-Dive Questions"):
        ws = wb[sheet]
        # Locate the answer column by HEADER. The template grew an "Example
        # Response" column, so the answer column is no longer D - writing to a
        # fixed index silently fills the examples and leaves the real column
        # empty, which the dump then reads as "0/56 answered".
        hdr = [str(ws.cell(2, c).value or "") for c in range(1, ws.max_column + 1)]
        col = next((i + 1 for i, h in enumerate(hdr)
                    if h in ("Customer Response", "Your Response")), None)
        if col is None:
            raise SystemExit(f"{sheet}: no response column in {hdr}")
        for r in range(3, ws.max_row + 1):
            ref = ws.cell(r, 1).value
            if isinstance(ref, str) and ref[:1] in "CD" and ref[1:].isdigit():
                v = ans.get(ref, "")
                if v:
                    ws.cell(r, col).value = v

    for sheet, rows in (("Appendix A - Accounts", appendix_a(p)),
                        ("Appendix B - IP Plan", appendix_b(p)),
                        ("Appendix C - Teams", appendix_c(p))):
        ws = wb[sheet]
        for i, row in enumerate(rows):
            for j, val in enumerate(row):
                ws.cell(4 + i, 1 + j).value = val

    ws = wb["Start Here"]
    ws.cell(2, 1).value = (f"Customer scenario: {p['name']} ({p['country'].upper()}), "
                           f"{p['industry']}, migrating from "
                           f"{p['source']}. Synthetic profile for validation.")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return sum(1 for v in ans.values() if v)


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("-o", "--out", default=str(REPO / "dist" / "customer-profiles-2"))
    args = ap.parse_args()
    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)

    blank = out / "_blank.xlsx"
    r = subprocess.run([sys.executable, "-X", "utf8", "-m",
                        "lz_pipeline.tools.gen_questionnaire", "-o", str(blank)],
                       capture_output=True, text=True, cwd=str(REPO),
                       env={**__import__("os").environ,
                            "PYTHONPATH": str(REPO / "pipeline")})
    if r.returncode != 0:
        print(r.stdout + r.stderr)
        return 1

    for i, p in enumerate(PROFILES, start=11):
        name = (f"HuaweiCloud-LZ-Assessment-{i}_{p['slug']}_"
                f"{p['country']}_{p['source']}.xlsx")
        n = fill(blank, out / name, p)
        print(f"  {name}  ({p['depth']}, {n}/56 answered)")
    blank.unlink()
    print(f"\n{len(PROFILES)} questionnaires -> {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
