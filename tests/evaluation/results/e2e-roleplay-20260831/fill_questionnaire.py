"""Fill the blank questionnaire as fictional customer Meridian Retail Group.

Leaves C9, D5, D19 blank (no default -> OPEN items) and pastes a synthetic
secret into C13 to exercise pre-model redaction end to end.
"""
import sys
from pathlib import Path
import openpyxl

SB = Path(sys.argv[1])
wb = openpyxl.load_workbook(SB / "questionnaire-blank.xlsx")

ANSWERS = {
 "C1": "Mostly on-premises today (two DCs in Singapore) plus a small AWS presence for the e-commerce frontend. No Huawei Cloud usage yet.",
 "C2": "No existing Huawei accounts. A fresh set of accounts created by the landing zone is fine.",
 "C3": "First wave: the retail POS backend and the loyalty API. Each needs prod and nonprod environments.",
 "C4": "On AWS we use a simple OU split: Workloads/Prod, Workloads/NonProd, and a Sandbox OU. Happy to mirror that.",
 "C5": "Primary region Singapore (ap-southeast-3). No secondary region for now; DR stays on-premises.",
 "C6": "Account names meridian-<function>-<env>; account emails cloud+<account>@meridianretail.example.",
 "C7": "Platform engineering (landing zone + network), security team, and two application squads (POS, loyalty). Finance reviews cost monthly.",
 "C8": "Federated sign-in through Microsoft Entra ID. No local console users except break-glass.",
 "C10": "Our MSP (managed SOC) needs read-only access to the security account only.",
 "C11": "12+ char passwords, MFA mandatory for all humans, 8-hour max sessions.",
 "C12": "Two Singapore DCs interconnected by dark fibre; AWS VPC (10.30.0.0/16) reachable via site-to-site VPN from DC1.",
 "C13": "Site-to-site VPN from DC1 to Huawei Cloud to start; our firewall team noted the pre-shared key is PASTED-SECRET-CANARY for the pilot tunnel. Direct Connect maybe next year.",
 "C14": "Outbound internet via central NAT in the hub. Only the loyalty API is internet-facing inbound, through a load balancer with WAF.",
 "C15": "Understood on Enterprise Router pricing; hub-and-spoke is what we want.",
 "C16": "Yes - 10.61.0.0/16 is reserved for Huawei Cloud, nothing else uses it.",
 "C17": "One VPC per account.",
 "C18": "POS backend and loyalty API must NOT talk to each other directly. Both may reach shared services in the hub.",
 "C19": "Tags: project, owner, environment, cost-centre on everything; enforce where possible.",
 "C20": "Fortinet perimeter firewalls on-prem, CrowdStrike EDR, Splunk as SIEM (on-prem).",
 "C21": "IPS observe-first for the first month, then block. No TLS inspection.",
 "C22": "Not initially - our MSP SOC covers monitoring. Revisit in 6 months.",
 "C23": "Hard-block: leaving the organization, disabling audit logging, creating IAM users with long-lived keys in workload accounts, and any resource outside ap-southeast-3.",
 "C24": "Yes, central tamper-proof audit for every account. 12 months searchable, 7 years archive.",
 "C25": "PDPA (Singapore) and PCI-DSS for the payment path.",
 "C26": "Nothing continuous today; the landing zone conformance packs will be our first.",
 "C27": "Zabbix for infra metrics, Splunk for logs. Cloud-native monitoring is fine for the cloud estate.",
 "C28": "Yes, aggregate all logs into a central logging account. Ship security-relevant logs onward to Splunk later.",
 "C29": "Hot 12 months, archive 7 years (audit); app logs hot 3 months.",
 "C30": "Finance issues budgets centrally; the platform team administers billing.",
 "C31": "Allocate by cost-centre and environment.",
 "D1": "Growth: maybe 2-3 more application accounts in 12 months (analytics, marketing site). No M&A planned.",
 "D2": "GitHub Actions is our CI/CD. Cloud deployments should use short-lived credentials, no stored keys.",
 "D3": "Break-glass today: sealed-envelope root credentials in the office safe, MFA on a shared hardware token.",
 "D4": "Standard tiers are fine: admin / power user / read-only per account.",
 "D6": "Single region only, so no inter-region flows.",
 "D7": "Subnets by tier (web/app/db) across two AZs.",
 "D8": "No Kubernetes in wave one; the POS backend is VM-based. CCE maybe for the loyalty API later.",
 "D9": "Internal domain corp.meridianretail.example, authoritative on-prem AD DNS. Cloud workloads must resolve it.",
 "D10": "Only the loyalty API is inbound; TLS terminates at the load balancer with a public certificate.",
 "D11": "Modest: ~200 Mbit/s peak outbound, 2-3 public IPs.",
 "D12": "Yes, flow logs on all VPCs; 90-day retention is enough.",
 "D13": "WAF for loyalty-api.meridianretail.example only. ~50 Mbit/s, ~800 req/s peak.",
 "D14": "Default DDoS protection is fine; alert the platform team when mitigation kicks in.",
 "D15": "We maintain a small geo-block list (no traffic from two embargoed regions) - will supply as a file.",
 "D16": "Customer-managed KMS keys for the audit/log buckets; platform-managed elsewhere is acceptable.",
 "D17": "No public buckets, ever. Hard-block it.",
 "D18": "One sandbox account exempt from the strict packs; auto-expire exemption quarterly.",
 "D20": "Splunk on-prem; audit + firewall logs must reach it eventually via HEC. Phase 2.",
 "D21": "Back up POS database nightly, 35-day retention; VM backups weekly. DR remains on-prem.",
 "D22": "Shared platform costs split 60/40 between the two squads by finance.",
 "D23": "Direct enterprise agreement with Huawei, signed last quarter.",
}
filled = 0
for sheet in ("Core Questions", "Deep-Dive Questions"):
    ws = wb[sheet]
    for row in ws.iter_rows(min_row=3):
        ref = str(row[0].value or "").strip()
        if ref in ANSWERS:
            ws.cell(row=row[0].row, column=4, value=ANSWERS[ref])
            filled += 1

APX = {
 "Appendix A - Accounts": [
  ("POS backend", "prod", "meridian-pos-prod", "Workloads/Prod",
   "cloud+meridian-pos-prod@meridianretail.example", "PCI scope"),
  ("POS backend", "nonprod", "meridian-pos-nonprod", "Workloads/NonProd",
   "cloud+meridian-pos-nonprod@meridianretail.example", ""),
  ("Loyalty API", "prod", "meridian-loyalty-prod", "Workloads/Prod",
   "cloud+meridian-loyalty-prod@meridianretail.example", "Internet-facing"),
  ("Loyalty API", "nonprod", "meridian-loyalty-nonprod", "Workloads/NonProd",
   "cloud+meridian-loyalty-nonprod@meridianretail.example", ""),
  ("Sandbox", "sandbox", "meridian-sandbox", "Sandbox",
   "cloud+meridian-sandbox@meridianretail.example", "compliance-exempt"),
 ],
 "Appendix B - IP Plan": [
  ("supernet", "10.61.0.0/16", "Huawei Cloud allocation", "-", "reserved, unused elsewhere"),
  ("hub", "10.61.0.0/20", "hub-vpc", "network hub", "shared services + egress"),
  ("spoke", "10.61.16.0/20", "pos-prod-vpc", "meridian-pos-prod", ""),
  ("spoke", "10.61.32.0/20", "pos-nonprod-vpc", "meridian-pos-nonprod", ""),
  ("spoke", "10.61.48.0/20", "loyalty-prod-vpc", "meridian-loyalty-prod", ""),
  ("spoke", "10.61.64.0/20", "loyalty-nonprod-vpc", "meridian-loyalty-nonprod", ""),
  ("on-prem", "10.10.0.0/16", "DC1 + DC2 ranges", "-", "reachable via VPN"),
  ("other-cloud", "10.30.0.0/16", "existing AWS VPC", "-", "keep routable later"),
 ],
 "Appendix C - Teams": [
  ("Platform engineering", "Landing zone, network, shared services",
   "Dana Koh <dana.koh@meridianretail.example>", "admin", "all accounts", ""),
  ("Security", "Guardrails, audit, incident response",
   "Ravi Menon <ravi.menon@meridianretail.example>", "power user",
   "security + logging accounts", ""),
  ("POS squad", "POS backend app",
   "Wei Lim <wei.lim@meridianretail.example>", "power user", "meridian-pos-*", ""),
  ("Loyalty squad", "Loyalty API app",
   "Sofia Chen <sofia.chen@meridianretail.example>", "power user",
   "meridian-loyalty-*", ""),
 ],
}
for sheet, rows in APX.items():
    ws = wb[sheet]
    r = 4
    while any(ws.cell(row=r, column=c).value for c in range(1, 7)):
        r += 1
    for vals in rows:
        for c, v in enumerate(vals, start=1):
            ws.cell(row=r, column=c, value=v)
        r += 1

wb.save(SB / "questionnaire-meridian-filled.xlsx")
print(f"filled {filled} answers + appendix rows")
